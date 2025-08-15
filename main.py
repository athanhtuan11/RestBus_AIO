# ==== E2E Project Config: Sinh và Load ==== 
import pandas as pd
import math
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import Autogen


# Biến toàn cục
WorkingPath = ''
DBCPath = ''
RestbusPath = ''
NodeName = ''
Channel = ''
e2e_type_map = {}
e2e_cfg_map = {}
global duplicate_signals
# Hàm phụ trợ

def get_doc_path():
    return os.path.join(os.path.expanduser('~'), 'Documents')

def save_log():
    doc_path = get_doc_path()
    log_path = os.path.join(doc_path, 'dblog')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write(f"{WorkingFolder_var.get()}\n{DBC_var.get()}\n{Node_var.get()}\n{Channel_var.get()}\n{tb_RestbusPath_var.get()}\n{E2EConfigPath_var.get()}\n")

def load_log():
    doc_path = get_doc_path()
    log_path = os.path.join(doc_path, 'dblog')
    if os.path.exists(log_path):
        with open(log_path, 'r', encoding='utf-8') as f:
            lines = f.read().splitlines()
            return lines + [''] * (6 - len(lines))
    return [''] * 6

def get_info():
    import pandas as pd
    global DBCPath, NodeName, WorkingPath
    DBCPath = DBC_var.get()
    NodeName = Node_var.get()
    path = WorkingFolder_var.get()
    if not path.endswith('\\'):
        path += '\\'
    WorkingPath = path

    # --- Xử lý file Excel tương tự GetInfo VBA nếu cần ---
    # Ví dụ: excel_path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files", "*.xlsx;*.xls")])
    # Nếu bạn muốn tự động lấy file theo WorkingPath, có thể sửa lại dòng dưới:
    excel_path = os.path.join(WorkingPath, "input.xlsx")
    if not os.path.isfile(excel_path):
        # Tạo file Excel mẫu nếu chưa có
        import pandas as pd
        columns = ["Command", "Data", "Value", "Expected", "Comment", "TestcaseID"]
        # Dữ liệu mẫu, bạn có thể chỉnh lại cho phù hợp
        data = [
            ["Comments", "Heading1", "", "", "REQ001", "TC001"],
            ["Set", "Var1", "1", "", "", ""],
            ["Check", "Var1", "", "[OK]", "", ""],
            ["Comments", "Heading2", "", "", "REQ002", "TC002"],
            ["Set", "Var2", "2", "", "", ""],
            ["Check", "Var2", "", "[PASS]", "", ""],
        ]
        df_sample = pd.DataFrame(data, columns=columns)
        df_sample.to_excel(excel_path, index=False)
    df = pd.read_excel(excel_path)
    SIOData = []
    ReqID = ""
    Index = -1
    CommandCol = "Command"
    DataCol = "Data"
    ValueCol = "Value"
    ExpectedCol = "Expected"
    CommentCol = "Comment"
    TestcaseIDCol = "TestcaseID"
    EndRow = len(df)
    new_flag = True
    new_id = 0
    if new_flag:
        Index += 1
        SIOData.append({
            "ParHeading": excel_path.split("\\")[-1].split(".")[0],
            "TestID": new_id,
            "Heading": "New",
            "HeadingFlag": True,
            "ReqID": "",
            "Input": "",
            "Expected": "",
            "Result": "",
        })
    for i in range(1, EndRow):
        row = df.iloc[i]
        if pd.notna(row[CommentCol]):
            ReqID = row[CommentCol]
        if row[CommandCol] == "Comments":
            Index += 1
            SIOData.append({
                "Heading": row[DataCol],
                "Result": "passed",
                "Input": "",
                "Expected": "",
                "ReqID": "",
                "TestID": "",
                "HeadingFlag": "",
                "ParHeading": "",
            })
            if pd.notna(df.iloc[0][TestcaseIDCol]):
                SIOData[Index]["ReqID"] = f"{ReqID}_{df.iloc[0][TestcaseIDCol]}"
            else:
                SIOData[Index]["ReqID"] = ReqID
        else:
            if "Set" in str(row[CommandCol]):
                SIOData[Index]["Input"] += f"|Set {row[DataCol]} = {str(row[ValueCol]).replace('[','').replace(']','')}"
            if "Check" in str(row[CommandCol]):
                SIOData[Index]["Expected"] += f"|{row[DataCol]} = {str(row[ExpectedCol])}"

def bt_curdir_click():
    WorkingFolder_var.set(os.getcwd())

def bt_opendbc_click():
    import subprocess
    path = DBC_var.get()
    if path:
        exe = r"C:\Program Files (x86)\Vector CANalyzer 8.1\Exec32\candb.exe"
        try:
            subprocess.Popen([exe, path])
        except FileNotFoundError:
            messagebox.showerror('Lỗi', f'Không tìm thấy file: {exe}')
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể mở DBC: {e}')
    else:
        messagebox.showwarning('Thông báo', 'Chưa chọn file DBC!')

def GetDBCInfor_Button_Click():
    get_info()
    GetDBCInfo(WorkingPath, DBCPath, NodeName)



def GenerateButton_click():
    import pandas as pd
    from tkinter import messagebox
    get_info()
    # Đọc file output DBC
    excel_path = os.path.join(WorkingPath, "DBC_output.xlsx")
    if not os.path.isfile(excel_path):
        messagebox.showerror("Lỗi", f"Không tìm thấy file {excel_path}, hãy chạy Get DBC Info trước!")
        return
    try:
        df = pd.read_excel(excel_path, sheet_name="Output_dbc")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không đọc được file {excel_path}: {e}")
        return
    node_name = Node_var.get().strip()
    channel = Channel_var.get().strip().upper()
    path = WorkingFolder_var.get().strip()
    if not path.endswith("\\"):
        path += "\\"
    if channel == "PRI":
        gen_path = os.path.join(path, "gen_PRI.can")
    elif channel == "PUB":
        gen_path = os.path.join(path, "gen_PUB.can")
    else:
        gen_path = os.path.join(path, "Gen.can")
    # Duyệt từng dòng, gom message TX có Node chứa NodeName, lấy chu kỳ đầu tiên, không lặp tên
    msg_dict = {}
    all_signals = []
    msg_signals_map = {}
    for idx, row in df.iterrows():
        if str(row.get("DIR", "")).strip() == "TX" and node_name in str(row.get("Node", "")):
            msg_name = str(row.get("Message", "")).strip()
            if msg_name:
                if msg_name not in msg_dict:
                    try:
                        cycle = int(float(row.get("Cyclic", 0))) if not pd.isna(row.get("Cyclic", 0)) else 0
                    except:
                        cycle = 0
                    msg_dict[msg_name] = cycle
                # Collect signals for duplicate detection
                sig = str(row.get("Signal", "")).strip()
                if msg_name not in msg_signals_map:
                    msg_signals_map[msg_name] = []
                if sig:
                    msg_signals_map[msg_name].append(sig)
                    all_signals.append(sig)
    # Detect duplicate signals
    from collections import Counter
    signal_counts = Counter(all_signals)
    duplicate_signals = {sig for sig, count in signal_counts.items() if count > 1}
    # Sắp xếp tên message
    msg_list = []
    for idx, (k, v) in enumerate(sorted(msg_dict.items(), key=lambda x: x[0])):
        msg_list.append({"Name": k, "Cycle": v, "Index": idx})
    # Tạo riêng 2 danh sách message cho PRI và PUB
    msg_dict_PRI = {}
    msg_dict_PUB = {}
    all_signals_PRI = []
    all_signals_PUB = []
    msg_signals_map_PRI = {}
    msg_signals_map_PUB = {}
    for idx, row in df.iterrows():
        if str(row.get("DIR", "")).strip() == "TX":
            msg_name = str(row.get("Message", "")).strip()
            node_val = str(row.get("Node", "")).upper()
            # PRI: Node chứa node_name
            if node_name in node_val:
                if msg_name:
                    if msg_name not in msg_dict_PRI:
                        try:
                            cycle = int(float(row.get("Cyclic", 0))) if not pd.isna(row.get("Cyclic", 0)) else 0
                        except:
                            cycle = 0
                        msg_dict_PRI[msg_name] = cycle
                    sig = str(row.get("Signal", "")).strip()
                    if msg_name not in msg_signals_map_PRI:
                        msg_signals_map_PRI[msg_name] = []
                    if sig:
                        msg_signals_map_PRI[msg_name].append(sig)
                        all_signals_PRI.append(sig)
            # PUB: Node chứa "PUB" hoặc channel là "PUB"
            if "PUB" in node_val or ("PUB" == channel):
                if msg_name:
                    if msg_name not in msg_dict_PUB:
                        try:
                            cycle = int(float(row.get("Cyclic", 0))) if not pd.isna(row.get("Cyclic", 0)) else 0
                        except:
                            cycle = 0
                        msg_dict_PUB[msg_name] = cycle
                    sig = str(row.get("Signal", "")).strip()
                    if msg_name not in msg_signals_map_PUB:
                        msg_signals_map_PUB[msg_name] = []
                    if sig:
                        msg_signals_map_PUB[msg_name].append(sig)
                        all_signals_PUB.append(sig)
    # Detect duplicate signals
    from collections import Counter
    signal_counts_PRI = Counter(all_signals_PRI)
    duplicate_signals_PRI = {sig for sig, count in signal_counts_PRI.items() if count > 1}
    signal_counts_PUB = Counter(all_signals_PUB)
    duplicate_signals_PUB = {sig for sig, count in signal_counts_PUB.items() if count > 1}
    # Sắp xếp tên message
    msg_list_PRI = []
    for idx, (k, v) in enumerate(sorted(msg_dict_PRI.items(), key=lambda x: x[0])):
        msg_list_PRI.append({"Name": k, "Cycle": v, "Index": idx})
    msg_list_PUB = []
    for idx, (k, v) in enumerate(sorted(msg_dict_PUB.items(), key=lambda x: x[0])):
        msg_list_PUB.append({"Name": k, "Cycle": v, "Index": idx})
    # Chọn msg_list theo channel hiện tại để sinh file .can
    msg_list = msg_list_PRI if channel == "PRI" else msg_list_PUB
    # Sinh file Gen.can
    try:
        with open(gen_path, "w", encoding="utf-8") as f:
            f.write("/*@!Encoding:1252*/\n")
            f.write("includes\n{\n\n}\n\n")
            f.write("variables\n{\n")
            f.write("\n\t//-----------------GEN: Define MSG Start----------------------\n\n")
            for msg in msg_list:
                if channel == "PUB":
                    f.write(f"\tmessage can1.{msg['Name']}\t{msg['Name']} = {{DIR = TX}};\n")
                else:
                    f.write(f"\tmessage can2.{msg['Name']}\t{msg['Name']} = {{DIR = TX}};\n")
            f.write("\n\t//-----------------GEN: Define MSG End------------------------\n\n")
            # MSG_CONFIG_TX_ENUM
            f.write("\n\t//----------------GEN: MSG ID End-------------------\n")
            f.write("\t// Generate MSG TX CONFIG\n")
            f.write("\tenum MSG_CONFIG_TX_ENUM {\n")
            f.write("\t/* 0 */ MSG_CONFIG_TX_ENUM_START = 0,\n")
            for i, msg in enumerate(msg_list):
                f.write(f"\t/* {i+1} */\tMSG_CONFIG_TX_{msg['Name']},\n")
            f.write(f"\t/* {len(msg_list)+1} */\tMSG_CONFIG_TX_ENUM_TX_END \n")
            f.write("\t};\n")
            f.write("\n\t//----------------GEN: MSG ID End-------------------\n\n")
            # Timer
            cycles = sorted(set([m["Cycle"] for m in msg_list if m["Cycle"] > 0]))
            f.write("\n\t//----------------GEN: Define Timer Start-------------------\n\n")
            min_cycle = None
            if cycles:
                min_cycle = cycles[0]
            tMinK1 = min_cycle
            used_cycles = set()
            while tMinK1 is not None:
                f.write(f"\tmsTimer PROJ_SendMSGTimer_{tMinK1}ms;\n")
                used_cycles.add(tMinK1)
                next_cycle = 10000
                for cyc in cycles:
                    if cyc > tMinK1 and cyc < next_cycle and cyc not in used_cycles:
                        next_cycle = cyc
                if next_cycle == 10000:
                    break
                tMinK1 = next_cycle
            f.write("\n\t//----------------GEN: Define Timer End-------------------\n\n")
            
            # Chỉ sinh bảng CRC_16_H1021_Tbl nếu channel là PRI
            if channel == "PRI":
                f.write("word CRC_16_H1021_Tbl[256] =\n{")
                crc16 = [
                    0x0000, 0x1021, 0x2042, 0x3063, 0x4084, 0x50a5, 0x60c6, 0x70e7,
                    0x8108, 0x9129, 0xa14a, 0xb16b, 0xc18c, 0xd1ad, 0xe1ce, 0xf1ef,
                    0x1231, 0x0210, 0x3273, 0x2252, 0x52b5, 0x4294, 0x72f7, 0x62d6,
                    0x9339, 0x8318, 0xb37b, 0xa35a, 0xd3bd, 0xc39c, 0xf3ff, 0xe3de,
                    0x2462, 0x3443, 0x0420, 0x1401, 0x64e6, 0x74c7, 0x44a4, 0x5485,
                    0xa56a, 0xb54b, 0x8528, 0x9509, 0xe5ee, 0xf5cf, 0xc5ac, 0xd58d,
                    0x3653, 0x2672, 0x1611, 0x0630, 0x76d7, 0x66f6, 0x5695, 0x46b4,
                    0xb75b, 0xa77a, 0x9719, 0x8738, 0xf7df, 0xe7fe, 0xd79d, 0xc7bc,
                    0x48c4, 0x58e5, 0x6886, 0x78a7, 0x0840, 0x1861, 0x2802, 0x3823,
                    0xc9cc, 0xd9ed, 0xe98e, 0xf9af, 0x8948, 0x9969, 0xa90a, 0xb92b,
                    0x5af5, 0x4ad4, 0x7ab7, 0x6a96, 0x1a71, 0x0a50, 0x3a33, 0x2a12,
                    0xdbfd, 0xcbdc, 0xfbbf, 0xeb9e, 0x9b79, 0x8b58, 0xbb3b, 0xab1a,
                    0x6ca6, 0x7c87, 0x4ce4, 0x5cc5, 0x2c22, 0x3c03, 0x0c60, 0x1c41,
                    0xedae, 0xfd8f, 0xcdec, 0xddcd, 0xad2a, 0xbd0b, 0x8d68, 0x9d49,
                    0x7e97, 0x6eb6, 0x5ed5, 0x4ef4, 0x3e13, 0x2e32, 0x1e51, 0x0e70,
                    0xff9f, 0xefbe, 0xdfdd, 0xcffc, 0xbf1b, 0xaf3a, 0x9f59, 0x8f78,
                    0x9188, 0x81a9, 0xb1ca, 0xa1eb, 0xd10c, 0xc12d, 0xf14e, 0xe16f,
                    0x1080, 0x00a1, 0x30c2, 0x20e3, 0x5004, 0x4025, 0x7046, 0x6067,
                    0x83b9, 0x9398, 0xa3fb, 0xb3da, 0xc33d, 0xd31c, 0xe37f, 0xf35e,
                    0x02b1, 0x1290, 0x22f3, 0x32d2, 0x4235, 0x5214, 0x6277, 0x7256,
                    0xb5ea, 0xa5cb, 0x95a8, 0x8589, 0xf56e, 0xe54f, 0xd52c, 0xc50d,
                    0x34e2, 0x24c3, 0x14a0, 0x0481, 0x7466, 0x6447, 0x5424, 0x4405,
                    0xa7db, 0xb7fa, 0x8799, 0x97b8, 0xe75f, 0xf77e, 0xc71d, 0xd73c,
                    0x26d3, 0x36f2, 0x0691, 0x16b0, 0x6657, 0x7676, 0x4615, 0x5634,
                    0xd94c, 0xc96d, 0xf90e, 0xe92f, 0x99c8, 0x89e9, 0xb98a, 0xa9ab,
                    0x5844, 0x4865, 0x7806, 0x6827, 0x18c0, 0x08e1, 0x3882, 0x28a3,
                    0xcb7d, 0xdb5c, 0xeb3f, 0xfb1e, 0x8bf9, 0x9bd8, 0xabbb, 0xbb9a,
                    0x4a75, 0x5a54, 0x6a37, 0x7a16, 0x0af1, 0x1ad0, 0x2ab3, 0x3a92,
                    0xfd2e, 0xed0f, 0xdd6c, 0xcd4d, 0xbdaa, 0xad8b, 0x9de8, 0x8dc9,
                    0x7c26, 0x6c07, 0x5c64, 0x4c45, 0x3ca2, 0x2c83, 0x1ce0, 0x0cc1,
                    0xef1f, 0xff3e, 0xcf5d, 0xdf7c, 0xaf9b, 0xbfba, 0x8fd9, 0x9ff8,
                    0x6e17, 0x7e36, 0x4e55, 0x5e74, 0x2e93, 0x3eb2, 0x0ed1, 0x1ef0
                ]
                for i in range(0, 256, 8):
                    f.write("\n    /*{:3d}:*/ ".format(i) + ', '.join(f"0x{v:04x}" for v in crc16[i:i+8]))
                    if i+8 < 256:
                        f.write(",")
                f.write("\n};\n\n")
            # Gọi hàm liệt kê PUB messages theo E2E type (nếu là PUB)
            if channel == "PUB":
                f.write("//----------------GEN: E2E CFG Start-------------------\n\n")
                e2e_list = get_pub_messages_e2e_info()
                f.write("const MSG_E2E_CFG_NULL=0;\n")
                f.write("const E2E_NONE = 0;\n")
                f.write("const E2E_CRC8 = 1;\n")
                f.write("const E2E_XOR = 2;\n")
                f.write("const E2E_CRC8_MULT = 3;\n\n")
                f.write("struct CAN_MSG_E2E_TX_ST\n{\n\tbyte PROJ_E2E_TYPE;\n\t byte PROJ_E2E_CFG;\n};\n\n")
                f.write("struct CAN_MSG_E2E_TX_ST PROJ_MSG_E2E_TX_ARR[MSG_CONFIG_TX_ENUM_TX_END +1] =\n{")
                f.write("\n/* 0  MSG_CONFIG_TX_ENUM_START*/    {E2E_NONE,   MSG_E2E_CFG_NULL  },\n")
                for idx, (msg_name, e2e_type, e2e_cfg) in enumerate(e2e_list, 1):
                    f.write(f"/* {idx}  MSG_CONFIG_TX_{msg_name} */    {{{e2e_type},   {e2e_cfg}  }},\n")
                f.write(f"/* {len(e2e_list)+1}  MSG_CONFIG_TX_ENUM_TX_END */    {{E2E_NONE,   MSG_E2E_CFG_NULL  }}\n")
                f.write("};\n\n")
                f.write("//----------------GEN: E2E CFG End-------------------\n")
            # Chèn bảng CRC_8_H1D_Tbl
            f.write("byte CRC_8_H1D_Tbl[256] =\n{")
            crc8 = [
                0x00, 0x1d, 0x3a, 0x27, 0x74, 0x69, 0x4e, 0x53,
                0xe8, 0xf5, 0xd2, 0xcf, 0x9c, 0x81, 0xa6, 0xbb,
                0xcd, 0xd0, 0xf7, 0xea, 0xb9, 0xa4, 0x83, 0x9e,
                0x25, 0x38, 0x1f, 0x02, 0x51, 0x4c, 0x6b, 0x76,
                0x87, 0x9a, 0xbd, 0xa0, 0xf3, 0xee, 0xc9, 0xd4,
                0x6f, 0x72, 0x55, 0x48, 0x1b, 0x06, 0x21, 0x3c,
                0x4a, 0x57, 0x70, 0x6d, 0x3e, 0x23, 0x04, 0x19,
                0xa2, 0xbf, 0x98, 0x85, 0xd6, 0xcb, 0xec, 0xf1,
                0x13, 0x0e, 0x29, 0x34, 0x67, 0x7a, 0x5d, 0x40,
                0xfb, 0xe6, 0xc1, 0xdc, 0x8f, 0x92, 0xb5, 0xa8,
                0xde, 0xc3, 0xe4, 0xf9, 0xaa, 0xb7, 0x90, 0x8d,
                0x36, 0x2b, 0x0c, 0x11, 0x42, 0x5f, 0x78, 0x65,
                0x94, 0x89, 0xae, 0xb3, 0xe0, 0xfd, 0xda, 0xc7,
                0x7c, 0x61, 0x46, 0x5b, 0x08, 0x15, 0x32, 0x2f,
                0x59, 0x44, 0x63, 0x7e, 0x2d, 0x30, 0x17, 0x0a,
                0xb1, 0xac, 0x8b, 0x96, 0xc5, 0xd8, 0xff, 0xe2,
                0x26, 0x3b, 0x1c, 0x01, 0x52, 0x4f, 0x68, 0x75,
                0xce, 0xd3, 0xf4, 0xe9, 0xba, 0xa7, 0x80, 0x9d,
                0xeb, 0xf6, 0xd1, 0xcc, 0x9f, 0x82, 0xa5, 0xb8,
                0x03, 0x1e, 0x39, 0x24, 0x77, 0x6a, 0x4d, 0x50,
                0xa1, 0xbc, 0x9b, 0x86, 0xd5, 0xc8, 0xef, 0xf2,
                0x49, 0x54, 0x73, 0x6e, 0x3d, 0x20, 0x07, 0x1a,
                0x6c, 0x71, 0x56, 0x4b, 0x18, 0x05, 0x22, 0x3f,
                0x84, 0x99, 0xbe, 0xa3, 0xf0, 0xed, 0xca, 0xd7,
                0x35, 0x28, 0x0f, 0x12, 0x41, 0x5c, 0x7b, 0x66,
                0xdd, 0xc0, 0xe7, 0xfa, 0xa9, 0xb4, 0x93, 0x8e,
                0xf8, 0xe5, 0xc2, 0xdf, 0x8c, 0x91, 0xb6, 0xab,
                0x10, 0x0d, 0x2a, 0x37, 0x64, 0x79, 0x5e, 0x43,
                0xb2, 0xaf, 0x88, 0x95, 0xc6, 0xdb, 0xfc, 0xe1,
                0x5a, 0x47, 0x60, 0x7d, 0x2e, 0x33, 0x14, 0x09,
                0x7f, 0x62, 0x45, 0x58, 0x0b, 0x16, 0x31, 0x2c,
                0x97, 0x8a, 0xad, 0xb0, 0xe3, 0xfe, 0xd9, 0xc4
            ]
            for i in range(0, 256, 8):
                f.write("\n    /*{:3d}:*/ ".format(i) + ', '.join(f"0x{v:02x}" for v in crc8[i:i+8]))
                if i+8 < 256:
                    f.write(",")
            f.write("\n};\n\n")
            f.write("}\n// varaiable End\n\n")
            
            # On Start
            f.write("ON Start\n{\n")
            f.write("\tPROJ_Init();\n")
            sysvar_type = channel  # Giữ lại biến này để không lỗi các hàm cũ, nhưng giá trị là channel (PUB/PRI)
            f.write(f"\tPROJ_Sim_{channel}_Signals_Init();\n")
            f.write("}\n\n")
            # PROJ_Init
            f.write("PROJ_Init ()\n{\n")
            f.write("\n\t//----------------GEN: Set Timer Start-------------------\n\t{\n")
            tMinK1 = min_cycle
            used_cycles = set()
            while tMinK1 is not None:
                f.write(f"\t\tsetTimer (PROJ_SendMSGTimer_{tMinK1}ms,  10);\n")
                used_cycles.add(tMinK1)
                next_cycle = 10000
                for cyc in cycles:
                    if cyc > tMinK1 and cyc < next_cycle and cyc not in used_cycles:
                        next_cycle = cyc
                if next_cycle == 10000:
                    break
                tMinK1 = next_cycle
            f.write("\t}\n\t//----------------GEN: Set Timer End-------------------\n\n}\n\n")

            # On Timer
            f.write("//----------------GEN: On Timer Start-------------------\n\n")
            # Nếu có message không có chu kỳ (cycle = 0), sinh block on key HOME
            
            msg_no_cycle = [msg for msg in msg_list if msg["Cycle"] == 0]
            if msg_no_cycle:
                f.write("on key HOME\n{\n\n\t//Calc and send MSG\n\t{\n")
                for msg in msg_no_cycle:
                    f.write(f"\t\tPROJ_CalcMsg_{msg['Name']}();\n\t\tSendMSG_{sysvar_type}(MSG_CONFIG_TX_{msg['Name']}, {msg['Name']});\n\n")
                f.write("\t}\n}\n")
            
            tMinK1 = min_cycle
            used_cycles = set()
            # Sinh các block on timer
            while tMinK1 is not None:
                f.write(f"on timer PROJ_SendMSGTimer_{tMinK1}ms\n{{\n\t//reset Timer\n\tsetTimer(PROJ_SendMSGTimer_{tMinK1}ms, {tMinK1});\n\n\t//Calc and send MSG\n\t{{\n")
                for i, msg in enumerate(msg_list):
                    if msg["Cycle"] == tMinK1:
                        f.write(f"\t\tPROJ_CalcMsg_{msg['Name']}();\n\t\tSendMSG_{sysvar_type}(MSG_CONFIG_TX_{msg['Name']}, {msg['Name']});\n\n")
                f.write("\t}\n}\n")
                used_cycles.add(tMinK1)
                next_cycle = 10000
                for cyc in cycles:
                    if cyc > tMinK1 and cyc < next_cycle and cyc not in used_cycles:
                        next_cycle = cyc
                if next_cycle == 10000:
                    break
                tMinK1 = next_cycle
            f.write("//----------------GEN: On Timer End-------------------\n")
            # Block SendMSG_
            f.write(f"\nSendMSG_{sysvar_type}(int MSGIndex, message* MSG)\n" + "{" + "\n\t\tPROJ_Sim_Timeout(MSGIndex, MSG);\n}\n")
            # Block PROJ_Sim_Timeout
            f.write(f"""PROJ_Sim_Timeout(int MSGIndex, message* MSG)
                {{
                    /* Timeout */
                    if (@Error_panel_{sysvar_type}::i_TOError[MSGIndex] != 0)
                    {{
                        if (@Error_panel_{sysvar_type}::i_TOErrorCounter[MSGIndex] > 0)
                        {{
                            @Error_panel_{sysvar_type}::i_TOErrorCounter[MSGIndex] --;
                        }}

                        if (@Error_panel_{sysvar_type}::i_TOErrorCounter[MSGIndex] == 0)
                        {{
                            @Error_panel_{sysvar_type}::i_TOError[MSGIndex] = 0;
                            @Error_panel_{sysvar_type}::i_TOErrorCounter[MSGIndex] = -1;
                        }}
                    
                    }}
                    else
                    {{
                        output(MSG);
                    }}
                }}
                """)
            # Block PROJ_Sim_<>_Signals_Init
            f.write(f"\nPROJ_Sim_{sysvar_type}_Signals_Init ()\n" + "{\n")
            for msg in msg_list:
                f.write(f"    /* {msg['Name']} */\n    {{\n")
                # Lấy tất cả signal thuộc message này và node hiện tại
                for idx, row in df.iterrows():
                    if str(row.get("DIR", "")).strip() == "TX" and str(row.get("Message", "")).strip() == msg['Name'] and node_name in str(row.get("Node", "")):
                        sig = str(row.get("Signal", "")).strip()
                        if sig:
                            f.write(f"        {msg['Name']}.{sig} = 0;\n")
                f.write("    }\n")
            f.write("}\n")
            # Block PROJ_Sim_Checksum
            selected_project = Project_var.get().strip()
            if selected_project == "Cherry":
                if channel == "PRI":
                    f.write("\nPROJ_Sim_Checksum(int MSGIndex, message* MSG)\n{")
                    f.write("\n\n\tword t_CRC_value;\n\tif ((MSGIndex == MSG_CONFIG_TX_RFC_DOPTraj)\n\t\t||(MSGIndex == MSG_CONFIG_TX_RFC_Fct)\n\t    ||(MSGIndex == MSG_CONFIG_TX_RFC_TIPLBoundary))\n\t\t\n\t{\n\t\tword data_id_t;\n\t\tbyte data_id[2];\n\t\tbyte data_temp[64];\n\t\tbyte index_t;\n\t\tbyte t_length;\n\t\tword t_CRC_value;   \n\t\tt_length = MSG.DataLength;\n\t\tdata_id_t = 4096 + MSG.id;\n\t\tfor (index_t=0;index_t<(t_length-2);index_t++)\n\t\t{\n\t\t\tdata_temp[index_t]= MSG.byte( index_t+2);\n\t\t}\n\t\tdata_temp[t_length -2] = data_id_t;\n\t\tdata_temp[t_length-1] = data_id_t>>8;\n\t\tt_CRC_value = Crc_CalculateCRC16_new(data_temp, t_length);\n\t\tif((@Error_panel_PRI::i_TOError[MSGIndex] ==0)\n\t\t&&(@Error_panel_PRI::i_CRC_CHKError[MSGIndex]))\n\t\t{\n\t\t\tif (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] > 0)\n\t\t\t{\n\t\t\t  @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] --;\n\t\t\t}\n\t\t\tif (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] == 0)\n\t\t\t{\n\t\t\t  @Error_panel_PRI::i_CRC_CHKError[MSGIndex] = 0;\n\t\t\t  @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] = -1;\n\t\t\t}\n\t\t\tt_CRC_value = t_CRC_value +1;\n\t\t}\n\t\telse\n\t\t{\n\t\t\t//Do nothing \n\t\t}\n\t\tMSG.byte(1) = t_CRC_value >>8;\n\t\tMSG.byte(0) = t_CRC_value;\n\t}\n\telse\n    {\n      t_CRC_value = PROJ_CHK_SAE_J1850_0x1D(MSG.DataLength, MSG);\n      if((@Error_panel_PRI::i_TOError[MSGIndex] ==0)\n          &&(@Error_panel_PRI::i_CRC_CHKError[MSGIndex]))\n      {\n          if (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] > 0)\n          {\n            @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] --;\n          }\n          if (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] == 0)\n          {\n            @Error_panel_PRI::i_CRC_CHKError[MSGIndex] = 0;\n            @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] = -1;\n          }\n          t_CRC_value = t_CRC_value +1;\n      }\n      else\n      {\n          //Do nothing \n      }\n      MSG.byte(0) = t_CRC_value;\n    }\n}\n")
                    # Thêm 2 hàm phụ trợ CRC
                    f.write("\nword Crc_CalculateCRC16_new(byte ptr[], byte p_length)\n{\n    byte    index;\n    word    crcTemp;\n    crcTemp = 0xFFFF;\n    for (index = 0; index < p_length; ++index)\n    {\n        crcTemp ^= ((word)ptr[index]) << 8;\n        crcTemp = (crcTemp << 8) ^ CRC_16_H1021_Tbl[(crcTemp >> 8) & (0xFF)];\n    }\n    return (crcTemp);\n}\n")
                    f.write("\nbyte PROJ_CHK_SAE_J1850_0x1D (byte p_length, message *data)\n{\n  byte t_RetCrc;\n  byte t_CRC_temp;\n  byte t_index;\n  t_CRC_temp = 0xFF;\n  for (t_index = 0; t_index < p_length-1; ++t_index)\n  {\n    t_CRC_temp ^= data.byte(t_index+1);\n    t_CRC_temp = CRC_8_H1D_Tbl[t_CRC_temp];\n  }\n  t_CRC_temp ^= 0xFF;\n  t_RetCrc = t_CRC_temp;\n  return t_RetCrc;\n}\n")
                else:
                    f.write("\nbyte PROJ_CHK_SAE_J1850_0x1D (byte p_length, message *data)\n{\n  byte t_RetCrc;\n  byte t_CRC_temp;\n  byte t_index;\n  t_CRC_temp = 0xFF;\n  for (t_index = 0; t_index < p_length-1; ++t_index)\n  {\n    t_CRC_temp ^= data.byte(t_index+1);\n    t_CRC_temp = CRC_8_H1D_Tbl[t_CRC_temp];\n  }\n  t_CRC_temp ^= 0xFF;\n  t_RetCrc = t_CRC_temp;\n  return t_RetCrc;\n}\n")
            else:  
                if channel == "PRI":
                    f.write("\nPROJ_Sim_Checksum(int MSGIndex, message* MSG)\n{")
                    f.write("\n\n\tword t_CRC_value;\n\tif ((MSGIndex == MSG_CONFIG_TX_RFC_DOPTraj)\n\t\t||(MSGIndex == MSG_CONFIG_TX_RFC_Fct)\n\t    ||(MSGIndex == MSG_CONFIG_TX_RFC_TIPLBoundary))\n\t\t\n\t{\n\t\tword data_id_t;\n\t\tbyte data_id[2];\n\t\tbyte data_temp[64];\n\t\tbyte index_t;\n\t\tbyte t_length;\n\t\tword t_CRC_value;   \n\t\tt_length = MSG.DataLength;\n\t\tdata_id_t = 4096 + MSG.id;\n\t\tfor (index_t=0;index_t<(t_length-2);index_t++)\n\t\t{\n\t\t\tdata_temp[index_t]= MSG.byte( index_t+2);\n\t\t}\n\t\tdata_temp[t_length -2] = data_id_t;\n\t\tdata_temp[t_length-1] = data_id_t>>8;\n\t\tt_CRC_value = Crc_CalculateCRC16_new(data_temp, t_length);\n\t\tif((@Error_panel_PRI::i_TOError[MSGIndex] ==0)\n\t\t&&(@Error_panel_PRI::i_CRC_CHKError[MSGIndex]))\n\t\t{\n\t\t\tif (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] > 0)\n\t\t\t{\n\t\t\t  @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] --;\n\t\t\t}\n\t\t\tif (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] == 0)\n\t\t\t{\n\t\t\t  @Error_panel_PRI::i_CRC_CHKError[MSGIndex] = 0;\n\t\t\t  @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] = -1;\n\t\t\t}\n\t\t\tt_CRC_value = t_CRC_value +1;\n\t\t}\n\t\telse\n\t\t{\n\t\t\t//Do nothing \n\t\t}\n\t\tMSG.byte(1) = t_CRC_value >>8;\n\t\tMSG.byte(0) = t_CRC_value;\n\t}\n\telse\n    {\n      t_CRC_value = PROJ_CHK_SAE_J1850_0x1D(MSG.DataLength, MSG);\n      if((@Error_panel_PRI::i_TOError[MSGIndex] ==0)\n          &&(@Error_panel_PRI::i_CRC_CHKError[MSGIndex]))\n      {\n          if (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] > 0)\n          {\n            @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] --;\n          }\n          if (@Error_panel_PRI::i_CHKErrorCounter[MSGIndex] == 0)\n          {\n            @Error_panel_PRI::i_CRC_CHKError[MSGIndex] = 0;\n            @Error_panel_PRI::i_CHKErrorCounter[MSGIndex] = -1;\n          }\n          t_CRC_value = t_CRC_value +1;\n      }\n      else\n      {\n          //Do nothing \n      }\n      MSG.byte(0) = t_CRC_value;\n    }\n}\n")
                    # Thêm 2 hàm phụ trợ CRC
                    f.write("\nword Crc_CalculateCRC16_new(byte ptr[], byte p_length)\n{\n    byte    index;\n    word    crcTemp;\n    crcTemp = 0xFFFF;\n    for (index = 0; index < p_length; ++index)\n    {\n        crcTemp ^= ((word)ptr[index]) << 8;\n        crcTemp = (crcTemp << 8) ^ CRC_16_H1021_Tbl[(crcTemp >> 8) & (0xFF)];\n    }\n    return (crcTemp);\n}\n")
                    f.write("\nbyte PROJ_CHK_SAE_J1850_0x1D (byte p_length, message *data)\n{\n  byte t_RetCrc;\n  byte t_CRC_temp;\n  byte t_index;\n  t_CRC_temp = 0xFF;\n  for (t_index = 0; t_index < p_length-1; ++t_index)\n  {\n    t_CRC_temp ^= data.byte(t_index+1);\n    t_CRC_temp = CRC_8_H1D_Tbl[t_CRC_temp];\n  }\n  t_CRC_temp ^= 0xFF;\n  t_RetCrc = t_CRC_temp;\n  return t_RetCrc;\n}\n")
                else:
                    f.write("\nbyte PROJ_CHK_SAE_J1850_0x1D (byte p_length, message *data)\n{\n  byte t_RetCrc;\n  byte t_CRC_temp;\n  byte t_index;\n  t_CRC_temp = 0xFF;\n  for (t_index = 0; t_index < p_length-1; ++t_index)\n  {\n    t_CRC_temp ^= data.byte(t_index+1);\n    t_CRC_temp = CRC_8_H1D_Tbl[t_CRC_temp];\n  }\n  t_CRC_temp ^= 0xFF;\n  t_RetCrc = t_CRC_temp;\n  return t_RetCrc;\n}\n")
            # Block PROJ_Sim_Rollingcounter
            selected_project = Project_var.get().strip()
            if selected_project == "Cherry":
                if channel == "PRI":
                    f.write("\nPROJ_Sim_Rollingcounter(int MSGIndex, message* MSG)\n{")
                    f.write("\n   if ((MSGIndex == MSG_CONFIG_TX_RFC_DOPTraj)\n\t\t||(MSGIndex == MSG_CONFIG_TX_RFC_Fct)\n\t    ||(MSGIndex == MSG_CONFIG_TX_RFC_TIPLBoundary))\n\t\n\t{\t\n\t\tbyte t_LastCounter;\n\t\tt_LastCounter = MSG.byte(2);   \n\t\t\n\t\t/* Rolling Counter */\n\t\tif ( (@Error_panel_PRI::i_TOError[MSGIndex] == 0)\n\t\t\t&& (@Error_panel_PRI::i_CRC_CHKError[MSGIndex] == 0)\n\t\t\t&& (@Error_panel_PRI::i_MCError[MSGIndex]))\n\t\t{\n\t\t\tif (@Error_panel_PRI::i_MCErrorCounter[MSGIndex] > 0)\n\t\t\t{\n\t\t\t\t@Error_panel_PRI::i_MCErrorCounter[MSGIndex] --;\n\t\t\t}\n\t\t\t\n\t\t\t//write(\"##RollingCounter_Error_Sim: : @CAN ID:0x%x @ReferNo.%d\", MSG.id, @Error_panel_PRI::i_MCErrorCounter[MSGIndex]);\n\t\t\t\n\t\t\tif (@Error_panel_PRI::i_MCErrorCounter[MSGIndex] ==0)\n\t\t\t{\n\t\t\t\t@Error_panel_PRI::i_MCError[MSGIndex] = 0;\n\t\t\t\t@Error_panel_PRI::i_MCErrorCounter[MSGIndex] = -1;\n\t\t\t}\n\t\t}\n\t\telse\n\t\t{\n\t\t\tt_LastCounter = (t_LastCounter +1) % 256;\n\t\t}\n\t\t\n\t\tMSG.byte(2) = t_LastCounter & 0xFF; \n\t}\n\telse\n    {\n        byte t_LastCounter;\n        byte t_rc_index;\n        t_rc_index = 1;\n        t_LastCounter = (MSG.byte(t_rc_index) & 0x0F); \n        \n        /* Rolling Counter */\n        if ( (@Error_panel_PRI::i_TOError[MSGIndex] == 0)\n            && (@Error_panel_PRI::i_CRC_CHKError[MSGIndex] == 0)\n            && (@Error_panel_PRI::i_MCError[MSGIndex]))\n        {\n            if (@Error_panel_PRI::i_MCErrorCounter[MSGIndex] > 0)\n            {\n                @Error_panel_PRI::i_MCErrorCounter[MSGIndex] --;\n            }\n            \n            \n            if (@Error_panel_PRI::i_MCErrorCounter[MSGIndex] ==0)\n            {\n                @Error_panel_PRI::i_MCError[MSGIndex] = 0;\n                @Error_panel_PRI::i_MCErrorCounter[MSGIndex] = -1;\n            }\n        }\n        else\n        {\n            t_LastCounter = (t_LastCounter +1) % 15;\n        }\n\n        MSG.byte(t_rc_index) = ((MSG.byte(t_rc_index) & 0xF0) | t_LastCounter); \n    }\n}\n")
                elif channel == "PUB":
                    # Sinh code C cho Rollingcounter PUB
                    f.write("\nPROJ_Sim_Rollingcounter(int MSGIndex, message* MSG)\n{")
                    f.write("\n    PROJ_Sim_Rollingcounter_XOR(MSGIndex, MSG);\n")
                    f.write("    PROJ_Sim_Rollingcounter_CRC8_MULT(MSGIndex, MSG);\n")
                    f.write("    PROJ_Sim_Rollingcounter_CRC8(MSGIndex, MSG);\n}\n")
                    # Sinh code C cho PROJ_Sim_Rollingcounter_XOR với đầy đủ logic
                    f.write("\nPROJ_Sim_Rollingcounter_XOR(int MSGIndex, message* MSG)\n{")
                    f.write("\n  if ((PROJ_MSG_E2E_TX_ARR[MSGIndex].PROJ_E2E_TYPE == E2E_XOR))\n  {\n    char t_LastSeqCounter; \n    byte t_LastSeqCounter_byte; \n    byte t_ALC_position;\n    byte t_ALC_Offset;\n\n    switch(MSGIndex)\n    {\n      case MSG_CONFIG_TX_ABS_ESP_1:\n              t_ALC_position = 7; \n              break;\n      case MSG_CONFIG_TX_ABS_ESP_3:\n      case MSG_CONFIG_TX_ABS_ESP_4:   \n      case MSG_CONFIG_TX_ABS_ESP_7:\n      case MSG_CONFIG_TX_ABS_ESP_8:  \n      case MSG_CONFIG_TX_BCM_1:\n      case MSG_CONFIG_TX_ICM_5:\n      case MSG_CONFIG_TX_TCU_2:\n      case MSG_CONFIG_TX_EMS_3:\n//    case MSG_CONFIG_TX_Z_FCM_FRM_6:\n              t_ALC_position = 6;   \n              break;\n      case MSG_CONFIG_TX_EMS_1:\n          t_ALC_position = 7;   \n          break;\n      case MSG_CONFIG_TX_YAS_1:\n      case MSG_CONFIG_TX_YAS_2:  \n              t_ALC_position = 0;   \n              break;\n//    case MSG_CONFIG_TX_Z_EPB_State_R:  \n//              t_ALC_position = 1;   \n//              break;\n      default: \n              t_ALC_position = 3;\n              break;\n    }\n    switch(MSGIndex)\n    {\n      case MSG_CONFIG_TX_ABS_ESP_1:\n      case MSG_CONFIG_TX_ABS_ESP_3:\n      case MSG_CONFIG_TX_ABS_ESP_4:   \n      case MSG_CONFIG_TX_ABS_ESP_7:\n      case MSG_CONFIG_TX_ABS_ESP_8:  \n      case MSG_CONFIG_TX_BCM_1: \n      case MSG_CONFIG_TX_EMS_1:\n      case MSG_CONFIG_TX_TCU_2:\n      case MSG_CONFIG_TX_ICM_5:\n//    case MSG_CONFIG_TX_Z_FCM_FRM_6:\n//    case MSG_CONFIG_TX_Z_EPB_State_R:\n              t_ALC_Offset = 0;  \n              break;\n      case MSG_CONFIG_TX_YAS_1:\n      case MSG_CONFIG_TX_YAS_2:  \n              t_ALC_Offset = 1;\n              break;\n      default: \n              t_ALC_Offset = 4;\n              break;\n    }\n    if (t_ALC_Offset == 0)\n    {\n      t_LastSeqCounter = (MSG.byte(t_ALC_position) & 0x0F); \n    }\n    else if (t_ALC_Offset == 4)\n    {\n      t_LastSeqCounter = (MSG.byte(t_ALC_position) & 0xF0) >>4; \n    }\n    else if (t_ALC_Offset == 1)\n    {\n    t_LastSeqCounter_byte = MSG.byte(0); \n    }\n    else\n    {\n     t_LastSeqCounter = (MSG.byte(t_ALC_position) & 0xF0) >>4;  \n    }\n    /* Rolling Counter */\n    if ( (@Error_panel_PUB::i_TOError[MSGIndex] == 0)\n        && (@Error_panel_PUB::i_CRC_CHKError[MSGIndex] == 0)\n        && (@Error_panel_PUB::i_MCError[MSGIndex]))\n    {\n        if (@Error_panel_PUB::i_MCErrorCounter[MSGIndex] > 0)\n        {\n            @Error_panel_PUB::i_MCErrorCounter[MSGIndex] --;\n        }\n        if (@Error_panel_PUB::i_MCErrorCounter[MSGIndex] ==0)\n        {\n            @Error_panel_PUB::i_MCError[MSGIndex] = 0;\n            @Error_panel_PUB::i_MCErrorCounter[MSGIndex] = -1;\n        }\n    }\n    else\n    {\n        t_LastSeqCounter = (t_LastSeqCounter +1) % 16;\n        t_LastSeqCounter_byte = (t_LastSeqCounter_byte +1) % 256;\n    }\n    if (t_ALC_Offset == 0)\n    {\n      MSG.byte(t_ALC_position) = (MSG.byte(t_ALC_position) & 0xF0) | (t_LastSeqCounter);  \n    }\n    else if (t_ALC_Offset == 4)\n    {\n      MSG.byte(t_ALC_position) = (MSG.byte(t_ALC_position) & 0x0F) | (t_LastSeqCounter <<4); \n    }\n    else if (t_ALC_Offset == 1)\n    {\n      MSG.byte(0) = (t_LastSeqCounter_byte & 0xFF) ; \n    }\n    else\n    {\n      MSG.byte(t_ALC_position) = (MSG.byte(t_ALC_position) & 0x0F) | (t_LastSeqCounter <<4); //ICM_1\n    }\n  }\n}\n")
                    # Sinh code C cho PROJ_Sim_Rollingcounter_CRC8_MULT
                    f.write("\nPROJ_Sim_Rollingcounter_CRC8_MULT(int MSGIndex, message* MSG)\n{")
                    f.write("\n  if ((PROJ_MSG_E2E_TX_ARR[MSGIndex].PROJ_E2E_TYPE == E2E_CRC8_MULT))\n  {\n    int msg_length;\n    int t_cycle;\n    int index;\n    msg_length = MSG.DataLength;\n    t_cycle=4;\n    for (index=0; index< t_cycle; index ++)\n    {\n      char t_LastSeqCounter; \n      byte t_ALC_position;\n      byte t_ALC_Offset;\n      t_ALC_position = 1 + 8*index;  \n      t_ALC_Offset = 0;\n      if (t_ALC_Offset == 0)\n      {\n        t_LastSeqCounter = (MSG.byte(t_ALC_position) & 0x0F); \n      }\n      else if (t_ALC_Offset == 4)\n      {\n        t_LastSeqCounter = (MSG.byte(t_ALC_position) & 0xF0) >>4; \n      }\n      /* Rolling Counter */\n      if ( (@Error_panel_PUB::i_TOError[MSGIndex] == 0)\n          && (@Error_panel_PUB::i_CRC_CHKError[MSGIndex] == 0)\n          && (@Error_panel_PUB::i_MCError[MSGIndex]))\n      {\n          if (@Error_panel_PUB::i_MCErrorCounter[MSGIndex] > 0)\n          {\n              @Error_panel_PUB::i_MCErrorCounter[MSGIndex] --;\n          }\n          if (@Error_panel_PUB::i_MCErrorCounter[MSGIndex] ==0)\n          {\n              @Error_panel_PUB::i_MCError[MSGIndex] = 0;\n              @Error_panel_PUB::i_MCErrorCounter[MSGIndex] = -1;\n          }\n      }\n      else\n      {\n          t_LastSeqCounter = (t_LastSeqCounter +1) % 16;\n      }\n      if (t_ALC_Offset == 0)\n      {\n        MSG.byte(t_ALC_position) = (MSG.byte(t_ALC_position) & 0xF0) | (t_LastSeqCounter);  \n      }\n      else if (t_ALC_Offset == 4)\n      {\n        MSG.byte(t_ALC_position) = (MSG.byte(t_ALC_position) & 0x0F) | (t_LastSeqCounter <<4); \n      }\n    }\n  }\n}\n")
                    # Sinh code C cho PROJ_Sim_Rollingcounter_CRC8
                    f.write("\nPROJ_Sim_Rollingcounter_CRC8(int MSGIndex, message* MSG)\n{")
                    f.write("\n  if ((PROJ_MSG_E2E_TX_ARR[MSGIndex].PROJ_E2E_TYPE == E2E_CRC8))\n  {\n    char t_LastSeqCounter; \n    byte t_ALC_position;\n    byte t_ALC_Offset;\n    switch(MSGIndex)\n    {\n      case MSG_CONFIG_TX_CLM_2:\n      case MSG_CONFIG_TX_EMS_6:\n              t_ALC_position = 2; \n              break;\n      case MSG_CONFIG_TX_SAM_1_G:\n              t_ALC_position = 5; \n              break;\n      case MSG_CONFIG_TX_EPS_2:\n      case MSG_CONFIG_TX_EPS_3: \n      case MSG_CONFIG_TX_EPS_4: \n      case MSG_CONFIG_TX_RLCR_1:  \n      case MSG_CONFIG_TX_RLCR_8:  \n      case MSG_CONFIG_TX_RLCR_9:  \n      case MSG_CONFIG_TX_RRCR_1:  \n      case MSG_CONFIG_TX_RRCR_7:  \n      case MSG_CONFIG_TX_RRCR_8:   \n      case MSG_CONFIG_TX_TCU_4:    \n      case MSG_CONFIG_TX_IPB_11:    \n      case MSG_CONFIG_TX_EMS_2_G: \n      case MSG_CONFIG_TX_DMS_1: \n      case MSG_CONFIG_TX_VCU_6: \n      case MSG_CONFIG_TX_HCU_17:\n      case MSG_CONFIG_TX_IHU_34:\n      case MSG_CONFIG_TX_IHU_35:\n//    case MSG_CONFIG_TX_Z_FCM_FRM_6:     \n              t_ALC_position = 6;   \n              break;\n      case MSG_CONFIG_TX_HCU_6:\n              t_ALC_position = 30;\n              break;\n      case MSG_CONFIG_TX_RLCR_3: \n      case MSG_CONFIG_TX_RLCR_4:  \n      case MSG_CONFIG_TX_RLCR_5:  \n      case MSG_CONFIG_TX_RLCR_6:  \n      case MSG_CONFIG_TX_RLCR_7:  \n      case MSG_CONFIG_TX_RRCR_2:  \n      case MSG_CONFIG_TX_RRCR_3:  \n      case MSG_CONFIG_TX_RRCR_4:  \n      case MSG_CONFIG_TX_RRCR_5:  \n      case MSG_CONFIG_TX_RRCR_6:   \n              t_ALC_position = 22;   \n              break;\n      default: \n              t_ALC_position = 1;\n              break;\n    }\n    switch(MSGIndex)\n    {\n      case MSG_CONFIG_TX_CLM_2:\n      case MSG_CONFIG_TX_MFS_2:\n      case MSG_CONFIG_TX_SAM_1_G:\n              t_ALC_Offset = 4;  \n              break;\n      case MSG_CONFIG_TX_EPS_2:\n      case MSG_CONFIG_TX_EPS_3:\n      case MSG_CONFIG_TX_EPS_4:\n      case MSG_CONFIG_TX_RLCR_3: \n      case MSG_CONFIG_TX_RLCR_4:  \n      case MSG_CONFIG_TX_RLCR_5:  \n      case MSG_CONFIG_TX_RLCR_6:  \n      case MSG_CONFIG_TX_RLCR_7:  \n      case MSG_CONFIG_TX_RRCR_2:  \n      case MSG_CONFIG_TX_RRCR_3:  \n      case MSG_CONFIG_TX_RRCR_4:  \n      case MSG_CONFIG_TX_RRCR_5:  \n      case MSG_CONFIG_TX_RRCR_6:\n      case MSG_CONFIG_TX_RLCR_1:  \n      case MSG_CONFIG_TX_RLCR_8:  \n      case MSG_CONFIG_TX_RLCR_9:  \n      case MSG_CONFIG_TX_RRCR_1:  \n      case MSG_CONFIG_TX_RRCR_7:  \n      case MSG_CONFIG_TX_RRCR_8:   \n      case MSG_CONFIG_TX_TCU_4:    \n      case MSG_CONFIG_TX_IPB_11: \n      case MSG_CONFIG_TX_EMS_6:    \n      case MSG_CONFIG_TX_EMS_2_G: \n      case MSG_CONFIG_TX_DMS_1: \n      case MSG_CONFIG_TX_VCU_6: \n      case MSG_CONFIG_TX_HCU_17:\n      case MSG_CONFIG_TX_IHU_34:\n      case MSG_CONFIG_TX_IHU_35:\n              t_ALC_Offset = 0;\n              break;\n      case MSG_CONFIG_TX_HCU_6:\n//    case MSG_CONFIG_TX_Z_FCM_FRM_6:\n      case MSG_CONFIG_TX_EPB_State_R:     \n              t_ALC_Offset = 0;\n              break;\n      default: \n              t_ALC_Offset = 1;\n              break;\n    }\n    if (t_ALC_Offset == 0)\n    {\n      t_LastSeqCounter = (MSG.byte(t_ALC_position) & 0x0F); \n    }\n    else if (t_ALC_Offset == 4)\n    {\n      t_LastSeqCounter = (MSG.byte(t_ALC_position) & 0xF0) >>4; \n    }\n    /* Rolling Counter */\n    if ( (@Error_panel_PUB::i_TOError[MSGIndex] == 0)\n        && (@Error_panel_PUB::i_CRC_CHKError[MSGIndex] == 0)\n        && (@Error_panel_PUB::i_MCError[MSGIndex]))\n    {\n        if (@Error_panel_PUB::i_MCErrorCounter[MSGIndex] > 0)\n        {\n            @Error_panel_PUB::i_MCErrorCounter[MSGIndex] --;\n        }\n        if (@Error_panel_PUB::i_MCErrorCounter[MSGIndex] ==0)\n        {\n            @Error_panel_PUB::i_MCError[MSGIndex] = 0;\n            @Error_panel_PUB::i_MCErrorCounter[MSGIndex] = -1;\n        }\n    }\n    else\n    {\n        t_LastSeqCounter = (t_LastSeqCounter +1) % 16;\n    }\n    if (t_ALC_Offset == 0)\n    {\n      MSG.byte(t_ALC_position) = (MSG.byte(t_ALC_position) & 0xF0) | (t_LastSeqCounter);  \n    }\n    else if (t_ALC_Offset == 4)\n    {\n      MSG.byte(t_ALC_position) = (MSG.byte(t_ALC_position) & 0x0F) | (t_LastSeqCounter <<4); \n    }\n  }\n}\n")
                    # Sinh code C cho Checksum PUB
                    f.write("\nPROJ_Sim_Checksum(int MSGIndex, message* MSG)\n{")
                    f.write("\n    PROJ_Sim_Checksum_XOR(MSGIndex, MSG);\n")
                    f.write("    PROJ_Sim_Checksum_CRC8_MULT(MSGIndex, MSG);\n")
                    f.write("    PROJ_Sim_Checksum_CRC8(MSGIndex, MSG);\n}\n")
                    # Sinh code C cho PROJ_Sim_Checksum_XOR (nội dung CRC8 như yêu cầu)
                    f.write("\nPROJ_Sim_Checksum_XOR(int MSGIndex, message* MSG)\n{")
                    f.write("\n  if ((PROJ_MSG_E2E_TX_ARR[MSGIndex].PROJ_E2E_TYPE == E2E_CRC8))\n  {\n    int msg_length;\n    int t_cycle;\n    int index;\n    word data1_id_t;\n    byte data1_id[2];\n    byte data1_temp[64];\n    byte index1_t;\n    byte t_length1;\n    word t_CRC16_value; \n    byte t_CRC8_value;\n    byte t_DataPtr[64];\n    byte t_length_ub;\n    byte i;\n    msg_length = MSG.DataLength;\n    if(MSGIndex== 39)\n    {\n      for (i=1;i< msg_length;i++)\n      {\n        t_DataPtr[i] = MSG.byte(i);\n      }\n    }\n    else if(MSGIndex== 15)\n    {\n      for (i=0;i< 3;i++)\n      {\n        t_DataPtr[i] = MSG.byte(i);\n      }\n      for (i=4;i< 8;i++)\n      {\n        t_DataPtr[i] = MSG.byte(i);\n      }\n    } \n    else \n    {\n       for (i=0;i< msg_length-1;i++)\n      {\n        t_DataPtr[i] = MSG.byte(i);\n      }\n    }\n    t_CRC8_value = PROJ_CHK_SAE_J1850_0x1D(msg_length-1, t_DataPtr);\n    if((@Error_panel_PUB::i_TOError[MSGIndex] ==0)\n          &&(@Error_panel_PUB::i_CRC_CHKError[MSGIndex]))\n      {\n          if (@Error_panel_PUB::i_CHKErrorCounter[MSGIndex] > 0)\n          {\n            @Error_panel_PUB::i_CHKErrorCounter[MSGIndex] --;\n          }\n          if (@Error_panel_PUB::i_CHKErrorCounter[MSGIndex] == 0)\n          {\n            @Error_panel_PUB::i_CRC_CHKError[MSGIndex] = 0;\n            @Error_panel_PUB::i_CHKErrorCounter[MSGIndex] = -1;\n          }\n          t_CRC8_value = t_CRC8_value +1;\n      }\n      else\n      {\n          //Do nothing \n      }\n      {\n        byte t_CRC_Position;\n        byte t_CRC_Offset;\n     switch(MSGIndex)\n    {\n      case MSG_CONFIG_TX_CLM_2:\n      case MSG_CONFIG_TX_EMS_2_G:\n      case MSG_CONFIG_TX_EPS_2:\n      case MSG_CONFIG_TX_EPS_3:\n      case MSG_CONFIG_TX_EPS_4:\n      case MSG_CONFIG_TX_RLCR_1:\n      case MSG_CONFIG_TX_RLCR_8:\n      case MSG_CONFIG_TX_RLCR_9:\n      case MSG_CONFIG_TX_RRCR_1:\n      case MSG_CONFIG_TX_RRCR_7:\n      case MSG_CONFIG_TX_RRCR_8:\n      case MSG_CONFIG_TX_TCU_4:\n      case MSG_CONFIG_TX_IPB_11:\n      case MSG_CONFIG_TX_DMS_1:\n      case MSG_CONFIG_TX_VCU_6:\n      case MSG_CONFIG_TX_HCU_17:\n      case MSG_CONFIG_TX_SAM_1_G:\n      case MSG_CONFIG_TX_IHU_34:\n      case MSG_CONFIG_TX_IHU_35:\n//    case MSG_CONFIG_TX_Z_FCM_FRM_6:\n              t_CRC_Position = 7;  \n              break;\n      case MSG_CONFIG_TX_HCU_6: \n              t_CRC_Position = 31;  \n              break;\n      case MSG_CONFIG_TX_RLCR_3: \n      case MSG_CONFIG_TX_RLCR_4:  \n      case MSG_CONFIG_TX_RLCR_5:  \n      case MSG_CONFIG_TX_RLCR_6:  \n      case MSG_CONFIG_TX_RLCR_7:  \n      case MSG_CONFIG_TX_RRCR_2:  \n      case MSG_CONFIG_TX_RRCR_3:  \n      case MSG_CONFIG_TX_RRCR_4:  \n      case MSG_CONFIG_TX_RRCR_5:  \n      case MSG_CONFIG_TX_RRCR_6: \n              t_CRC_Position = 23;\n              break;  \n      case MSG_CONFIG_TX_EMS_6: \n              t_CRC_Position = 3;\n              break;\n      default: \n              t_CRC_Position = 0;\n              break;\n    }\n      MSG.byte(t_CRC_Position) = t_CRC8_value;\n      } \n    }\n}\n")
                    # Sinh code C cho PROJ_Sim_Checksum_CRC8_MULT
                    f.write("\nPROJ_Sim_Checksum_CRC8_MULT(int MSGIndex, message* MSG)\n{")
                    f.write("\n    int msg_length;\n    int t_cycle;\n    int index;\n    word data1_id_t;\n    byte data1_id[2];\n    byte data1_temp[64];\n    byte index1_t;\n    byte t_length1;\n    msg_length = MSG.DataLength;\n    t_cycle = 4;//????VCU_2 ????crc\n    for (index=0; index< t_cycle; index ++)\n    {\n      byte t_CRC_value;\n      byte t_DataPtr[64];\n      byte t_length_ub;\n      byte i;\n \n      t_length_ub = MSG.DataLength;\n      for (i=1;i< 8;i++)\n      {\n        t_DataPtr[i] = MSG.byte((index*8)+i);\n      }\n      \n      \n      t_CRC_value = PROJ_CHK_SAE_J1850_0x1D(7, t_DataPtr);\n \n \t \n      \n      if((@Error_panel_PUB::i_TOError[MSGIndex] ==0)\n          &&(@Error_panel_PUB::i_CRC_CHKError[MSGIndex]))\n      {\n          if (@Error_panel_PUB::i_CHKErrorCounter[MSGIndex] > 0)\n          {\n            @Error_panel_PUB::i_CHKErrorCounter[MSGIndex] --;\n          }\n\n          if (@Error_panel_PUB::i_CHKErrorCounter[MSGIndex] == 0)\n          {\n            @Error_panel_PUB::i_CRC_CHKError[MSGIndex] = 0;\n            @Error_panel_PUB::i_CHKErrorCounter[MSGIndex] = -1;\n          }\n\n          t_CRC_value = t_CRC_value +1;\n      }\n      else\n      {\n          //Do nothing \n      }\n      \n      {\n        byte t_CRC_Position;\n        byte t_CRC_Offset;\n        \n        t_CRC_Position = 8*index;\n      \n        MSG.byte(t_CRC_Position) = t_CRC_value;\n  \t}\n  \n\n      \n    }\n}\n")
                    # Sinh code C cho PROJ_Sim_Checksum_CRC8
                    f.write("\nPROJ_Sim_Checksum_CRC8(int MSGIndex, message* MSG)\n{")
                    f.write("\n  else if ((PROJ_MSG_E2E_TX_ARR[MSGIndex].PROJ_E2E_TYPE == E2E_XOR)&MSGIndex!= 26)\n  {\n\n    byte t_CRC8_value;\n    byte t_CRC_position;\n    byte t_CRC_flag;\n    switch(MSGIndex)\n    {  \n      case MSG_CONFIG_TX_ABS_ESP_3:\n      case MSG_CONFIG_TX_ABS_ESP_4:   \n      case MSG_CONFIG_TX_ABS_ESP_7:\n      case MSG_CONFIG_TX_ABS_ESP_8:   \n      case MSG_CONFIG_TX_ICM_5:\n      case MSG_CONFIG_TX_TCU_2:\n      case MSG_CONFIG_TX_EMS_3:\n              t_CRC_flag=0;\n              t_CRC_position = 7; \n              break;\n      case MSG_CONFIG_TX_ABS_ESP_1:\n              t_CRC_flag=1;  \n              t_CRC_position = 5;   \n              break;\n      case MSG_CONFIG_TX_BCM_1: \n              t_CRC_flag=2;\n              t_CRC_position = 1;   \n              break;\n      case MSG_CONFIG_TX_YAS_1: \n              t_CRC_flag=3;\n              t_CRC_position = 7;   \n              break;\n      case MSG_CONFIG_TX_YAS_2: \n              t_CRC_flag=4;\n              t_CRC_position = 7; \n              break;\n      case MSG_CONFIG_TX_EMS_1: \n              t_CRC_flag=5;\n              t_CRC_position = 2;    \n              break;\n      default: \n              t_CRC_flag=5;\n              t_CRC_position = 0;\n              break;\n    }\n  if(t_CRC_flag == 0)\n  {\n\n    t_CRC8_value=((MSG.byte(0)+MSG.byte(1)+MSG.byte(2)+MSG.byte(3)+MSG.byte(4)+MSG.byte(5)+MSG.byte(6))&0xFF)^0xFF;\n  }\n  else if (t_CRC_flag == 1)\n  {\n    t_CRC8_value=((MSG.byte(0)+MSG.byte(1)+MSG.byte(2)+MSG.byte(3)+MSG.byte(4)+MSG.byte(6)+MSG.byte(7))&0xFF)^0xFF;\n  }\n  else if (t_CRC_flag == 2)\n  {\n    t_CRC8_value=((MSG.byte(0)+MSG.byte(2)+MSG.byte(3)+MSG.byte(4)+MSG.byte(5)+MSG.byte(6)+MSG.byte(7))&0xFF)^0xFF;\n  }\n  else if (t_CRC_flag == 3)\n  {\n   t_CRC8_value=0x01^0x27^MSG.byte(0)^MSG.byte(1) ^MSG.byte(2) ^MSG.byte(3) ^MSG.byte(4) ^MSG.byte(5) ^MSG.byte(6); \n  }\n  else if (t_CRC_flag == 4)\n  {\n   t_CRC8_value=0x01^0x28^MSG.byte(0)^MSG.byte(1) ^MSG.byte(2) ^MSG.byte(3) ^MSG.byte(4) ^MSG.byte(5) ^MSG.byte(6); \n  }\n  else if (t_CRC_flag == 5)\n  {\n   t_CRC8_value=((MSG.byte(0)+MSG.byte(1)+MSG.byte(3)+MSG.byte(4)+MSG.byte(5)+MSG.byte(6)+MSG.byte(7))&0xFF)^0xFF;\n  }\n  else\n  {\n  }\n          \n  if((@Error_panel_PUB::i_TOError[MSGIndex] ==0)\n        &&(@Error_panel_PUB::i_CRC_CHKError[MSGIndex]))\n    {\n        if (@Error_panel_PUB::i_CHKErrorCounter[MSGIndex] > 0)\n        {\n          @Error_panel_PUB::i_CHKErrorCounter[MSGIndex] --;\n        }\n\n        if (@Error_panel_PUB::i_CHKErrorCounter[MSGIndex] == 0)\n        {\n          @Error_panel_PUB::i_CRC_CHKError[MSGIndex] = 0;\n          @Error_panel_PUB::i_CHKErrorCounter[MSGIndex] = -1;\n        }\n\n        t_CRC8_value = t_CRC8_value +1;\n    }\n    else\n    {\n        //Do nothing \n    }\n    \n    {\n    \n      MSG.byte(t_CRC_position) = t_CRC8_value;\n\n    } \n  }\n}\n")
            else:
                f.write("\nPROJ_Sim_Rollingcounter(int MSGIndex, message* MSG)\n{")
                f.write("}\n")
                f.write("\nPROJ_Sim_Checksum(int MSGIndex, message* MSG)\n{")
                f.write("}\n")

            f.write(f"//--------------------------------------------------\n\n")
            f.write(f"//--------------------------------------------------\n")
            # Block PROJ_CalcMsg_<message> cho từng message TX
            node_var = Node_var.get().strip()
            for msg in msg_list:
                f.write(f"\nPROJ_CalcMsg_{msg['Name']}()\n" + "{\n")
                # Lấy tất cả signal thuộc message này và node hiện tại
                for idx, row in df.iterrows():
                    if str(row.get("DIR", "")).strip() == "TX" and str(row.get("Message", "")).strip() == msg['Name'] and node_name in str(row.get("Node", "")):
                        sig = str(row.get("Signal", "")).strip()
                        if not sig:
                            continue
                        # Đổi tên biến nếu trùng signal
                        if sig in duplicate_signals:
                            var_base = f"{sig}_{msg['Name']}"
                        else:
                            var_base = sig
                        f.write(f"    /* {sig} */\n")
                        f.write(f"    if(!@{node_var}::i_{var_base}_SIM)\n    {{\n")
                        f.write(f"        if(@{node_var}::i_{var_base}_EN)\n        {{\n")
                        f.write(f"            {msg['Name']}.{sig} = @{node_var}::i_{var_base}_Raw;\n")
                        f.write(f"        }}\n        else\n        {{\n")
                        f.write(f"            {msg['Name']}.{sig}.phys = @{node_var}::i_{var_base};\n")
                        f.write(f"        }}\n    }}\n    else\n    {{\n")
                        f.write(f"        // To be completed in simulation\n    }}\n")
                # Rolling counter, Checksum, DLC Check
                f.write(f"    /* Rolling counter */\n    PROJ_Sim_Rollingcounter(MSG_CONFIG_TX_{msg['Name']}, {msg['Name']});\n\n")
                f.write(f"    /* CheckSum */\n    PROJ_Sim_Checksum(MSG_CONFIG_TX_{msg['Name']}, {msg['Name']});\n\n")
                # Xác định sysvar_type cho lỗi DLC
                err_panel = get_err_panel_by_channel(channel)
                f.write(f"    /* DLC Check */\n    {msg['Name']}.DataLength = @Error_panel_{err_panel}::i_DLCError[MSG_CONFIG_TX_{msg['Name']}]?1:DBLookup({msg['Name']}).dlc;\n")
                f.write("}\n")
            if channel == "PRI":
                f.write("on message CAN1.*\n")
                f.write("{\n")
                f.write("    output(this);\n")
                f.write("}\n")
                f.write("on message CAN2.*\n")
                f.write("{\n")
                f.write("    output(this);\n")
                f.write("}\n")

            # Gọi GenVariableSignals sau khi sinh Gen.can
            node_var = Node_var.get().strip()
            GenVariableSignals(path, channel, node_var, msg_list)
            # Gọi Gen Main Panel
            GenMainPanel(path, channel, msg_list, main_form_name="MainPanel")
            # Gọi Gen Error Panel đúng theo channel hiện tại
            if channel == "PRI":
                GenErrorPanel(path, "PRI", msg_list_PRI)
            elif channel == "PUB":
                GenErrorPanel(path, "PUB", msg_list_PUB)
        messagebox.showinfo("Thông báo", f"Đã Gen Thành Công")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể ghi file {gen_path}: {e}")
# Helper to map channel to err_panel
def get_err_panel_by_channel(channel):
    channel = str(channel).upper()
    return channel

# --- Khung hàm GetDBCInfo để bạn phát triển tiếp ---
def filter_unwanted_signals(dbc_info):
    """
    Trả về tuple (kept, removed) với kept là list các dòng giữ lại, removed là list các dòng bị loại bỏ.
    Tùy chỉnh logic filter tại đây.
    Ví dụ: loại bỏ signal có tên chứa 'RESERVED' hoặc message có tên trong blacklist.
    """
    keywords = ["CRC", "CHECKSUM", "CHKSM", "ALIVECTR", "BLOCKCTR" ,"MSGCTR" , "MESSAGECOUNTER", "ALIVECOUNTER", "ROLLGCNTR" , "MSGCNTR", "MSGCOUNTER","MESSAGE_COUNTER","ALIVE","DIAG" , "XCP"]
    def is_remove(row):
        # 1. DIR là RX
        if str(row.get("DIR", "")).strip().upper() == "RX":
            return True
        # 2. Tên signal chứa các từ khóa (không phân biệt hoa thường)
        sig = str(row.get("Signal", "")).upper()
        for kw in keywords:
            if kw in sig:
                return True
        return False
    kept = []
    removed = []
    for row in dbc_info:
        if is_remove(row):
            removed.append(row)
        else:
            kept.append(row)
    return kept, removed

def GetDBCInfo(working_folder, dbc_path, node_name):
    """
    Đọc file DBC, phân tích nội dung và xử lý dữ liệu tương tự VBA GetDBCInfo.
    working_folder: thư mục làm việc
    dbc_path: đường dẫn file DBC
    node_name: tên node
    """
    import re
    import pandas as pd
    from tkinter import simpledialog
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        load_workbook = None
        Workbook = None
        PatternFill = None

    if not os.path.isfile(dbc_path):
        messagebox.showerror("Lỗi", f"Không tìm thấy file DBC: {dbc_path}")
        return
    with open(dbc_path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.read().splitlines()

    # Parse DBC
    dbc_info = []
    init_info = {}
    cycle_info = {}
    desc_info = {}
    bo_pattern = re.compile(r'^BO_ (\d+) (\w+):.* (\w+)$')
    sg_pattern = re.compile(r'^\s*SG_\s+(\w+)\s*:\s*(\d+)\|(\d+)@(\d+)([+-])\s*\(([-\d.Ee]+),([\-\d.Ee]+)\)\s*\[([\-\d.Ee]+)\|([\-\d.Ee]+)\]\s*"(.*?)"\s*(.*)$')
    ba_init_pattern = re.compile(r'^BA_ "GenSigStartValue".+SG_ \d+ (\w+) (-?[\d\.]+);')
    ba_cycle_pattern = re.compile(r'^BA_ "GenMsgCycleTime".+BO_ (\d+) (\d+);')
    val_pattern = re.compile(r'^VAL_ \d+ (\w+) (.+);')
    current_msg = None
    current_msgid = None
    current_node = None
    # Lấy biến Channel toàn cục
    global Channel
    # Hỏi người dùng có muốn xuất all node không (giống VBA)
    from tkinter import messagebox
    gen_all_node = messagebox.askyesno("Gen All Node?", "Do you want to get all Node?")
    # Lấy danh sách node từ BU_ đầu file
    bu_nodes = []
    for line in lines:
        if line.strip().startswith('BU_:'):
            bu_nodes = line.strip().split('BU_:')[1].strip().split()
            break
    for i, line in enumerate(lines):
        bo_match = bo_pattern.match(line)
        if bo_match:
            current_msgid = bo_match.group(1)
            current_msg = bo_match.group(2)
            current_node = bo_match.group(3)
            continue
        sg_match = sg_pattern.match(line)
        if sg_match and current_msg is not None:
            signal = sg_match.group(1)
            startbit = sg_match.group(2)
            length = sg_match.group(3)
            factor = sg_match.group(6)
            offset = sg_match.group(7)
            minv = sg_match.group(8)
            maxv = sg_match.group(9)
            unit = sg_match.group(10)
            node_field = sg_match.group(11)
            node_list = [n.strip() for n in node_field.split(',')] if node_field else ['']
            if gen_all_node:
                # Xuất cho tất cả node trong BU_, RX nếu current_node==node, TX ngược lại
                for node in bu_nodes:
                    if current_node == node:
                        direction = "RX"
                    else:
                        direction = "TX"
                    dbc_info.append({
                        "Node": node,
                        "Message": current_msg,
                        "MSGID": current_msgid,
                        "DIR": direction,
                        "Signal": signal,
                        "StartBit": startbit,
                        "Length": length,
                        "Factor": factor,
                        "Offset": offset,
                        "MinValue": minv,
                        "MaxValue": maxv,
                        "Unit": unit,
                        "Description": "",
                        "InitValue": 0,
                        "Cyclic": "",
                        "Norm": "",
                        "ConvFact": "",
                        "ALV/CHK_Flag": ""
                    })
            else:
                # Chỉ xuất signal cho node_name: TX nếu node_name là transmitter, RX nếu node_name là receiver
                if current_node == node_name:
                    # Nếu là transmitter, chỉ xuất RX
                    dbc_info.append({
                        "Node": node_name,
                        "Message": current_msg,
                        "MSGID": current_msgid,
                        "DIR": "RX",
                        "Signal": signal,
                        "StartBit": startbit,
                        "Length": length,
                        "Factor": factor,
                        "Offset": offset,
                        "MinValue": minv,
                        "MaxValue": maxv,
                        "Unit": unit,
                        "Description": "",
                        "InitValue": 0,
                        "Cyclic": "",
                        "Norm": "",
                        "ConvFact": "",
                        "ALV/CHK_Flag": ""
                    })
                elif node_name in node_list:
                    # Nếu là receiver (và không phải transmitter), chỉ xuất TX
                    dbc_info.append({
                        "Node": node_name,
                        "Message": current_msg,
                        "MSGID": current_msgid,
                        "DIR": "TX",
                        "Signal": signal,
                        "StartBit": startbit,
                        "Length": length,
                        "Factor": factor,
                        "Offset": offset,
                        "MinValue": minv,
                        "MaxValue": maxv,
                        "Unit": unit,
                        "Description": "",
                        "InitValue": 0,
                        "Cyclic": "",
                        "Norm": "",
                        "ConvFact": "",
                        "ALV/CHK_Flag": ""
                    })
            continue
        # Nếu không match regex, log ra console để debug
        if line.strip().startswith('SG_'):
            print('Không parse được:', line)
        ba_init_match = ba_init_pattern.match(line)
        if ba_init_match:
            init_info[ba_init_match.group(1)] = ba_init_match.group(2)
            continue
        ba_cycle_match = ba_cycle_pattern.match(line)
        if ba_cycle_match:
            cycle_info[ba_cycle_match.group(1)] = ba_cycle_match.group(2)
            continue
        val_match = val_pattern.match(line)
        if val_match:
            desc_info[val_match.group(1)] = val_match.group(2)
            continue
    # Gán init, cyclic, desc cho từng signal
    for row in dbc_info:
        sig = row["Signal"]
        msgid = row["MSGID"]
        if sig in init_info:
            try:
                row["InitValue"] = float(init_info[sig]) * float(row["Factor"]) + float(row["Offset"])
            except Exception:
                row["InitValue"] = init_info[sig]
        if msgid in cycle_info:
            row["Cyclic"] = cycle_info[msgid]
        if sig in desc_info:
            row["Description"] = desc_info[sig]
    # Sắp xếp theo Message, Signal
    dbc_info = sorted(dbc_info, key=lambda x: (x["Message"], x["Signal"]))
    # Không lọc node, xuất toàn bộ signal/message
    dbc_info = [row for row in dbc_info if row["Signal"]]

    # Lọc unwanted signals/messages, lưu lại sheet remove
    kept, removed = filter_unwanted_signals(dbc_info)

    out_path = os.path.join(working_folder, "DBC_output.xlsx")
    columns = ["Node", "Message", "DIR", "Cyclic", "Signal", "Length", "StartBit", "MinValue", "MaxValue", "Factor", "Offset", "InitValue", "Unit", "Description", "Norm", "ConvFact", "ALV/CHK_Flag"]
    import pandas as pd
    # Xóa file cũ nếu có
    if os.path.isfile(out_path):
        try:
            os.remove(out_path)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xóa file {out_path}. Hãy đóng file Excel trước khi chạy lại!\n{e}")
            return
    df_keep = pd.DataFrame(kept, columns=columns)
    df_remove = pd.DataFrame(removed, columns=columns)
    # Kiểm tra file có đang bị mở không
    try:
        if load_workbook is not None:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Output_dbc"
            for cidx, col in enumerate(columns, 1):
                ws1.cell(row=1, column=cidx, value=col)
            for ridx, row in enumerate(df_keep.values, 2):
                for cidx, val in enumerate(row, 1):
                    ws1.cell(row=ridx, column=cidx, value=val)
            # Tô màu xen kẽ cho Output_dbc
            if PatternFill:
                fill1 = PatternFill(start_color="BBD3DB", end_color="BBD3DB", fill_type="solid")
                fill2 = PatternFill(start_color="C7DF7B", end_color="C7DF7B", fill_type="solid")
                last_msg = None
                flag = True
                for ridx in range(2, len(df_keep) + 2):
                    msg = ws1.cell(row=ridx, column=2).value
                    if msg != last_msg:
                        flag = not flag
                        last_msg = msg
                    for cidx in range(1, len(columns) + 1):
                        ws1.cell(row=ridx, column=cidx).fill = fill1 if flag else fill2
            # Sheet remove
            ws2 = wb.create_sheet("remove")
            for cidx, col in enumerate(columns, 1):
                ws2.cell(row=1, column=cidx, value=col)
            for ridx, row in enumerate(df_remove.values, 2):
                for cidx, val in enumerate(row, 1):
                    ws2.cell(row=ridx, column=cidx, value=val)
            wb.save(out_path)
        else:
            # Ghi bằng pandas, sheet Output_dbc và remove
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_keep.to_excel(writer, sheet_name="Output_dbc", index=False)
                df_remove.to_excel(writer, sheet_name="remove", index=False)
    except PermissionError:
        messagebox.showerror("Lỗi", f"Không thể ghi file {out_path}. Hãy đóng file Excel trước khi chạy lại!")
        return
    messagebox.showinfo("Thông báo", f"Đã xuất dữ liệu ra file: {out_path} (sheet Output_dbc, remove)")

def commandbutton3_click():
    # SMUDoorsGen() # Placeholder
    #root.withdraw()
    def SMUDoorsGen():
        # TODO: Cập nhật code xử lý SMU Doors Gen ở đây
        pass

def commandbutton4_click():
    # SIODoorsGen() # Placeholder
    #root.withdraw()
    def SIODoorsGen():
        # TODO: Cập nhật code xử lý SIO Doors Gen ở đây
        pass

def channel_change(*args):
    global Channel
    Channel = Channel_var.get()
    save_log()

def commandbutton5_click():
    # GenDataBase() # Placeholder
    #root.withdraw()
    def GenDataBase():
        # TODO: Cập nhật code xử lý Gen DataBase ở đây
        pass

def commandbutton6_click():
    # GetSystemDegradation() # Placeholder
    #root.withdraw()
    def GetSystemDegradation():
        # TODO: Cập nhật code xử lý Get System Degradation ở đây
        pass

def commandbutton7_click():
    # GetFaultID() # Placeholder
    #root.withdraw()
    def GetFaultID():
        # TODO: Cập nhật code xử lý Get Fault ID ở đây
        pass

def dbc_change(*args):
    global DBCPath
    DBCPath = DBC_var.get()
    save_log()


def geta2l_click():
    # GetAtoL() # Placeholder
    #root.withdraw()
    def GetAtoL():
        # TODO: Cập nhật code xử lý Get A2L ở đây
        pass

def node_change(*args):
    global NodeName
    NodeName = Node_var.get()
    save_log()

def tb_RestbusPath_change(*args):
    global RestbusPath
    RestbusPath = tb_RestbusPath_var.get()
    save_log()

def workingfolder_change(*args):
    global WorkingPath
    WorkingPath = WorkingFolder_var.get()
    save_log()


# ==== GenVariableSignals: Sinh file Variable_signals_<sysvar_type>.xml (theo logic VBA) ====
def GenVariableSignals(path, sysvar_type, node_var, msg_list):

    import pandas as pd
    var_path = os.path.join(path, f"Variable_signals_{node_var}.xml")
    # Ghi header XML
    with open(var_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0" encoding="utf-8"?>\n')
        f.write('<systemvariables version="4">\n')
        f.write(' <namespace comment="" name="">\n')
        f.write(f'  <namespace comment="" name="{node_var}">\n')

    # Đọc dữ liệu Output_dbc để truyền vào GenPanel
    excel_path = os.path.join(path, "DBC_output.xlsx")
    if not os.path.isfile(excel_path):
        return
    try:
        df = pd.read_excel(excel_path, sheet_name="Output_dbc")
    except Exception:
        return

    # Lấy channel từ sysvar_type
    channel = sysvar_type.upper()

    # Xây dựng tập hợp tất cả tên signal xuất hiện nhiều hơn 1 lần trên toàn bộ msg_list
    from collections import Counter
    all_signals = []
    for msg in msg_list:
        msg_name = msg['Name']
        # Lấy signals thuộc message này
        signals = []
        nrow = len(df)
        i = 0
        while i < nrow:
            row = df.iloc[i]
            t_MsgName = str(row.get("Message", ""))
            if t_MsgName == msg_name:
                while i < nrow and str(df.iloc[i].get("Message", "")) == msg_name:
                    row2 = df.iloc[i]
                    sig = str(row2.get("Signal", ""))
                    if sig:
                        all_signals.append(sig)
                    i += 1
                break
            i += 1
    signal_counts = Counter(all_signals)
    duplicate_signals = {sig for sig, count in signal_counts.items() if count > 1}

    # Gọi GenPanel cho từng message, truyền vào tập hợp duplicate_signals
    for msg in msg_list:
        GenPanel(msg['Name'], path, channel, df, var_path, sysvar_type, duplicate_signals)

    # Đóng namespace
    with open(var_path, "a", encoding="utf-8") as f:
        f.write('   </namespace>\n')
        f.write('  </namespace>\n')
        f.write('</systemvariables>\n')
# Nút mở file DBC_output.xlsx
def open_dbc_output():
    path = WorkingFolder_var.get().strip()
    if not path.endswith("\\"):
        path += "\\"
    excel_path = os.path.join(path, "DBC_output.xlsx")
    if os.path.isfile(excel_path):
        try:
            os.startfile(excel_path)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở file: {e}")
    else:
        messagebox.showwarning("Thông báo", f"Không tìm thấy file: {excel_path}")

# ==== GenPanel: Sinh biến cho từng message trong XML và sinh file panel .xvp ====
def GenPanel(msg_name, path, channel, df, var_path, sysvar_type, duplicate_signals=None):

    # Lấy signals thuộc message này, đúng thứ tự và block như VBA
    signals = []
    SigCol = "Signal"
    NodeCol = "Node"
    LengthCol = "Length"
    MinCol = "MinValue"
    MaxCol = "MaxValue"
    InitCol = "InitValue"
    FactCol = "Factor"
    DescCol = "Description"
    ALVCHKCol = "ALV/CHK_Flag"
    UnitCol = "Unit"
    nrow = len(df)
    i = 0
    while i < nrow:
        row = df.iloc[i]
        t_MsgName = str(row.get("Message", ""))
        if t_MsgName == msg_name:
            # Bắt đầu block signals cho message này
            while i < nrow and str(df.iloc[i].get("Message", "")) == msg_name:
                row2 = df.iloc[i]
                sig = str(row2.get(SigCol, ""))
                node = str(row2.get(NodeCol, ""))
                alvchk = str(row2.get(ALVCHKCol, "")).lower()
                sig_lc = sig.lower()
                # Loại bỏ các signal không hợp lệ
                if (
                    "vector__xxx" in node.lower()
                    or "msgc" in sig_lc
                    or "alivec" in sig_lc
                    or "checksum" in sig_lc
                    or "chks" in sig_lc
                    or "blockcnt" in sig_lc
                    or "alv" in alvchk
                    or "chk" in alvchk
                ):
                    i += 1
                    continue
                if sig:
                    signals.append({
                        "Signal": sig,
                        "length": row2.get(LengthCol, ""),
                        "Min": row2.get(MinCol, ""),
                        "max": row2.get(MaxCol, ""),
                        "InitValue": row2.get(InitCol, ""),
                        "factor": row2.get(FactCol, ""),
                        "Description": row2.get(DescCol, ""),
                        "Unit": row2.get(UnitCol, "")
                    })
                i += 1
            break  # Chỉ lấy block đầu tiên của message này
        i += 1
    if not signals:
        return
    n = len(signals)
    NoHeight = int(math.sqrt(n))
    NoWidth = int(math.ceil(n / NoHeight)) if NoHeight else n
    if NoHeight * NoWidth < n:
        NoWidth += 1
    # Tạo file panel .xvp
    panel_path = os.path.join(path, f"{msg_name}.xvp")
    with open(panel_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0"?>\n')
        f.write('<Panel Type="Vector.CANalyzer.Panels.PanelSerializer, Vector.CANalyzer.Panels.Serializer, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null">\n')
        f.write(f' <Object Type="Vector.CANalyzer.Panels.Runtime.Panel, Vector.CANalyzer.Panels.Common, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="Panel" Children="Controls" ControlName="{msg_name}">\n')
        for i, sig in enumerate(signals):
            try:
                factor = float(sig["factor"])
            except Exception:
                factor = 1
            tType = "float" if factor < 1 else "int"
            ValueDisp = "Double" if factor < 1 else "Decimal"
            desc = str(sig["Description"])
            length = int(sig["length"]) if str(sig["length"]).isdigit() else 8
            minv = str(sig["Min"])
            maxv = str(sig["max"])
            base_name = sig["Signal"]
            # Đổi tên nếu trùng signal
            if duplicate_signals and base_name in duplicate_signals:
                name = f"{base_name}_{msg_name}"
            else:
                name = base_name
            location = f"{20 + 260 * (i % NoWidth)}, {20 + (20 + 80 + 20 + 10) * (i // NoWidth)}"
            f.write(f'  <Object Type="Vector.CANalyzer.Panels.Design.GroupBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="GroupBoxControl{i}" Children="Controls" ControlName="Group Box">\n')
            node_var = Node_var.get().strip()
            if re.search(r"\-\>.+\-\>", desc) and length <= 4:
                f.write(ComboBox(str(i), location, name, minv, maxv, node_var))
            else:
                f.write(Trackbar(str(i), location, name, minv, maxv, ValueDisp, node_var))
            f.write(Textbox_phys(str(i), f"20, 55", name, minv, maxv, ValueDisp, node_var))
            f.write(Textbox_hex(str(i), f"20, 80", name, minv, maxv, node_var))
            f.write(Checkbox_R(str(i), f"180, 80", name, minv, maxv, node_var))
            f.write(Checkbox_S(str(i), f"180, 55", name, minv, maxv, node_var))
            f.write(f'  <Property Name="Name">GroupBoxControl{i}</Property>\n')
            f.write(f'  <Property Name="Location">{location}</Property>\n')
            f.write(f'  <Property Name="Size">240, {20 + 80 + 20 + 5}</Property>\n')
            f.write(f'  <Property Name="Text">{name}</Property>\n')
            f.write(f'  </Object>\n')
        f.write(f'  <Property Name="Text">{msg_name}</Property>\n')
        f.write(f'  <Property Name="Size">{20 + 260 * NoWidth}, {20 + (80 + 20 + 20 + 10) * NoHeight}</Property>\n')
        f.write(' </Object>\n')
        f.write('</Panel>\n')

    # Ghi biến vào file Variable_signals_<sysvar_type>.xml đúng logic VBA
    node_var = Node_var.get().strip()
    # Nếu signal nằm trong duplicate_signals thì luôn thêm _<message name>
    with open(var_path, "a", encoding="utf-8") as fvar:
        for sig in signals:
            base_name = sig["Signal"]
            name = base_name
            if duplicate_signals and base_name in duplicate_signals:
                name = f"{base_name}_{msg_name}"
            init = sig["InitValue"]
            unit = sig["Unit"]
            desc = sig["Description"]
            tType = "float" if (float(sig["factor"]) if str(sig["factor"]).replace('.','',1).isdigit() else 1) < 1 else "int"
            length = sig["length"]
            fvar.write(CreateVariable(name, init, unit, desc, tType, length))
# ==== Các hàm sinh control XML cho panel (chuẩn VBA) ====
def Trackbar(ID, Location, Name, Min, max, ValueDisp, SysVar):
    node_var = Node_var.get().strip()
    tstr = f'   <Object Type="Vector.CANalyzer.Panels.Design.TrackBarControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="TrackBarControl{ID}" Children="Controls" ControlName="Track Bar">\n'
    tstr += f'    <Property Name="Name">TrackBarControl{ID}</Property>\n'
    tstr += f'    <Property Name="Size">200, 30</Property>\n'
    tstr += f'    <Property Name="Location">{Location}</Property>\n'
    if ValueDisp == "Double":
        tstr += f'    <Property Name="SmallChangeDouble">0.1</Property>\n'
    if ValueDisp == "Double":
        tstr += f'    <Property Name="MaximumDouble">{max}</Property>\n'
        tstr += f'    <Property Name="MinimumDouble">{Min}</Property>\n'
    else:
        tstr += f'    <Property Name="Maximum">{max}</Property>\n'
        tstr += f'    <Property Name="Minimum">{Min}</Property>\n'
    tstr += f'    <Property Name="SymbolConfiguration">4;16;{node_var}_Signals;;;i_{Name};2;;;-1</Property>\n'
    tstr += f'   </Object>\n'
    return tstr

def ComboBox(ID, Location, Name, Min, max, SysVar):
    node_var = Node_var.get().strip()
    tstr = f'   <Object Type="Vector.CANalyzer.Panels.Design.ComboBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="ComboBoxControl{ID}" Children="Controls" ControlName="Combo Box">\n'
    tstr += f'    <Property Name="Name">ComboBoxControl{ID}</Property>\n'
    tstr += f'    <Property Name="Size">200, 30</Property>\n'
    tstr += f'    <Property Name="Location">{Location}</Property>\n'
    tstr += f'    <Property Name="SymbolConfiguration">4;16;{node_var}_Signals;;;i_{Name};2;;;-1</Property>\n'
    tstr += f'   </Object>\n'
    return tstr

def Textbox_phys(ID, Location, Name, Min, max, ValueDisp, SysVar):
    node_var = Node_var.get().strip()
    tstr = f'   <Object Type="Vector.CANalyzer.Panels.Design.TextBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="TextBoxPhysControl{ID}" Children="Controls" ControlName="Input/Output Box">\n'
    tstr += f'    <Property Name="Name">TextBoxPhysControl{ID}</Property>\n'
    tstr += f'    <Property Name="Size">150, 20</Property>\n'
    tstr += f'    <Property Name="Location">{Location}</Property>\n'
    tstr += f'    <Property Name="ValueDisplay">{ValueDisp}</Property>\n'
    if ValueDisp == "Double":
        tstr += f'    <Property Name="ValueDecimalPlaces">1</Property>\n'
    tstr += f'    <Property Name="DescriptionText">Phys</Property>\n'
    tstr += f'    <Property Name="DescriptionSize">50, 20</Property>\n'
    tstr += f'    <Property Name="SymbolConfiguration">4;16;{node_var}_Signals;;;i_{Name};2;;;-1</Property>\n'
    tstr += f'   </Object>\n'
    return tstr

def Textbox_hex(ID, Location, Name, Min, max, SysVar):
    node_var = Node_var.get().strip()
    tstr = f'   <Object Type="Vector.CANalyzer.Panels.Design.TextBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="TextBoxControl{ID}" Children="Controls" ControlName="Input/Output Box">\n'
    tstr += f'    <Property Name="Name">TextBoxControl{ID}</Property>\n'
    tstr += f'    <Property Name="Size">150, 20</Property>\n'
    tstr += f'    <Property Name="Location">{Location}</Property>\n'
    tstr += f'    <Property Name="ValueDisplay">Hexadecimal</Property>\n'
    tstr += f'    <Property Name="DescriptionText">Hex</Property>\n'
    tstr += f'    <Property Name="DescriptionSize">50, 20</Property>\n'
    tstr += f'    <Property Name="SymbolConfiguration">4;16;{node_var}_Signals;;;i_{Name}_Raw;1;;;-1</Property>\n'
    tstr += f'   </Object>\n'
    return tstr

def Checkbox_R(ID, Location, Name, Min, max, SysVar):
    node_var = Node_var.get().strip()
    tstr = f'   <Object Type="Vector.CANalyzer.Panels.Design.CheckBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="CheckBoxControlR{ID}" Children="Controls" ControlName="Check Box">\n'
    tstr += f'    <Property Name="Name">CheckBoxControlR{ID}</Property>\n'
    tstr += f'    <Property Name="Size">55, 20</Property>\n'
    tstr += f'    <Property Name="Location">{Location}</Property>\n'
    tstr += f'    <Property Name="Text">RAW</Property>\n'
    tstr += f'    <Property Name="DescriptionText"></Property>\n'
    tstr += f'    <Property Name="SymbolConfiguration">4;16;{node_var}_Signals;;;i_{Name}_EN;1;;;-1</Property>\n'
    tstr += f'   </Object>\n'
    return tstr

def Checkbox_S(ID, Location, Name, Min, max, SysVar):
    node_var = Node_var.get().strip()
    tstr = f'   <Object Type="Vector.CANalyzer.Panels.Design.CheckBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="CheckBoxControlS{ID}" Children="Controls" ControlName="Check Box">\n'
    tstr += f'    <Property Name="Name">CheckBoxControlS{ID}</Property>\n'
    tstr += f'    <Property Name="Size">55, 20</Property>\n'
    tstr += f'    <Property Name="Location">{Location}</Property>\n'
    tstr += f'    <Property Name="Text">SIM</Property>\n'
    tstr += f'    <Property Name="DescriptionText"></Property>\n'
    tstr += f'    <Property Name="SymbolConfiguration">4;16;{node_var}_Signals;;;i_{Name}_SIM;1;;;-1</Property>\n'
    tstr += f'   </Object>\n'
    return tstr

def CreateVariable(Name, Init, Unit, Description, tType, tLength):
    # Format Init based on tType
    try:
        if tType == "int":
            # If Init is float but integer value, cast to int
            if isinstance(Init, str) and Init.strip() != "":
                val = float(Init)
                if val.is_integer():
                    Init_fmt = str(int(val))
                else:
                    Init_fmt = str(int(round(val)))
            else:
                Init_fmt = str(int(float(Init)))
        elif tType == "float":
            val = float(Init)
            if val == 0:
                Init_fmt = "0"
            else:
                Init_fmt = f"{val:.2f}"
        else:
            Init_fmt = str(Init)
    except Exception:
        Init_fmt = str(Init)

    tstr = f'      <variable name="i_{Name}" comment="" anlyzLocal="2" type="{tType}" startValue="{Init_fmt}" readOnly="false" valueSequence="false" unit="{Unit}">\n'
    desc_str = str(Description)
    # Support both VBA style (0->MOVING|1->STATIONARY|...) and Excel style (0 "MOVING" 1 "STATIONARY" ...)
    if tLength and int(tLength) <= 4:
        # Check for Excel style: value "desc" value "desc" ...
        import re
        excel_valdesc = re.findall(r'(\d+)\s+"([^"]+)"', desc_str)
        if excel_valdesc:
            tstr += f'       <valuetable definesMinMax="false">\n'
            for val, desc in excel_valdesc:
                tstr += f'        <valuetableentry value="{val}" description="{desc}" />\n'
            tstr += f'       </valuetable>\n'
        elif "->" in desc_str:
            tstr += f'       <valuetable definesMinMax="false">\n'
            for entry in desc_str.split("|"):
                if "->" in entry:
                    parts = entry.split("->")
                    tstr += f'        <valuetableentry value="{parts[0].strip()}" description="{parts[1].strip()}" />\n'
            tstr += f'       </valuetable>\n'
    tstr += f'      </variable>\n'
    tstr += f'      <variable name="i_{Name}_Raw" comment="" anlyzLocal="2" type="int" startValue="0" readOnly="false" valueSequence="false" unit="{Unit}">\n      </variable>\n'
    tstr += f'      <variable name="i_{Name}_EN" comment="" anlyzLocal="2" type="int" startValue="0" readOnly="false" valueSequence="false" unit="{Unit}">\n      </variable>\n'
    tstr += f'      <variable name="i_{Name}_SIM" comment="" anlyzLocal="2" type="int" startValue="1" readOnly="false" valueSequence="false" unit="{Unit}">\n      </variable>\n'
    return tstr

# ==== Gen Main Panel: Sinh file MSG_Panel_<sysvar_type>.xvp (chuẩn VBA) ====
def GenMainPanel(path, sysvar_type, msg_list, main_form_name="MainPanel"):

    main_panel_path = os.path.join(path, f"MSG_Panel_{sysvar_type}.xvp")
    # Tìm maxLen tên message
    names = [msg['Name'] for msg in msg_list if msg.get('Name')]
    if not names:
        return
    maxLen = max(len(name) for name in names)
    maxWidth = 8 * maxLen
    n = len(names)
    NoHeight = int(n ** 0.5)
    NoWidth = int(n / NoHeight) if NoHeight else n
    if NoHeight * NoWidth < n:
        NoWidth += 1
    # Điều chỉnh lại theo VBA
    NoWidth = int(1000 / maxWidth) if maxWidth else 1
    if NoWidth == 0:
        NoWidth = 1
    with open(main_panel_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0"?>\n')
        f.write('<Panel Type="Vector.CANalyzer.Panels.PanelSerializer, Vector.CANalyzer.Panels.Serializer, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null">\n')
        f.write(f' <Object Type="Vector.CANalyzer.Panels.Runtime.Panel, Vector.CANalyzer.Panels.Common, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="Panel" Children="Controls" ControlName="{main_form_name}">\n')
        for i, name in enumerate(names):
            x = 20 + (maxWidth + 10) * (i % NoWidth)
            y = 20 + 35 * (i // NoWidth)
            f.write(PanelButtonControl(str(maxWidth), name, f"{x}, {y}") + "\n")
        f.write(f'  <Property Name="Text">{main_form_name}</Property>\n')
        # Tính lại NoHeight cho Size
        NoHeight = n // NoWidth
        if n % NoWidth:
            NoHeight += 1
        f.write(f'  <Property Name="Size">{20 + (maxWidth + 10) * NoWidth}, {20 + 35 * NoHeight}</Property>\n')
        f.write(' </Object>\n')
        f.write('</Panel>\n')

# Hàm PanelButtonControl chuẩn VBA (Python)
def PanelButtonControl(width, msg, location):
    # Remove _ from msg for Name property
    name_no_underscore = msg.replace('_', '')
    tstr = f'   <Object Type="Vector.CANalyzer.Panels.Design.PanelButtonControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="PanelButtonControl{name_no_underscore}" Children="Controls" ControlName="Panel Control Button">\n'
    tstr += f'    <Property Name="Name">PanelButtonControl{name_no_underscore}</Property>\n'
    tstr += f'    <Property Name="Size">{width}, 25</Property>\n'
    tstr += f'    <Property Name="Location">{location}</Property>\n'
    tstr += f'    <Property Name="PanelList">1;1;{msg}.xvp</Property>\n'
    tstr += f'    <Property Name="Text">{msg}</Property>\n'
    tstr += f'   </Object>'
    return tstr
# ==== Gen Error Panel: Sinh file Error_Panel_<sysvar_type>.xvp (chuẩn VBA) ====
def GenErrorPanel(path, sysvar_type, msg_list):

    error_path = os.path.join(path, f"Error_Panel_{sysvar_type}.xvp")
    error_form_name = f"Error_Panel_{sysvar_type}"
    NoWidth = 4
    n = len(msg_list)
    NoHeight = n // NoWidth
    if NoHeight * NoWidth < n:
        NoHeight += 1
    with open(error_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0"?>\n')
        f.write('<Panel Type="Vector.CANalyzer.Panels.PanelSerializer, Vector.CANalyzer.Panels.Serializer, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null">\n')
        f.write(f' <Object Type="Vector.CANalyzer.Panels.Runtime.Panel, Vector.CANalyzer.Panels.Common, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="Panel" Children="Controls" ControlName="{error_form_name}">\n')
        for i, msg in enumerate(msg_list):
            name = msg.get('Name')
            if not name:
                continue
            idx1 = msg.get('Index', i) + 1
            f.write(f'  <Object Type="Vector.CANalyzer.Panels.Design.GroupBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="GroupBoxControl{i}" Children="Controls" ControlName="Group Box">\n')
            for j, errtype in enumerate(["TO", "CHK", "MC", "DLC"]):
                x = 10 + j * 50
                f.write(ErrCheckbox(name, f"{x}, 15", errtype, idx1, sysvar_type) + "\n")
                f.write(ErrTextbox(name, f"{x}, 35", errtype, idx1, sysvar_type) + "\n")
            f.write(f'  <Property Name="Name">GroupBoxControl{i}</Property>\n')
            f.write(f'  <Property Name="Location">{10 + 240 * (i % NoWidth)}, {20 + 60 * (i // NoWidth)}</Property>\n')
            f.write(f'  <Property Name="Size">230, 60</Property>\n')
            f.write(f'  <Property Name="Text">{name}</Property>\n')
            f.write(f'  </Object>\n')
        f.write(f'  <Property Name="Text">{error_form_name}</Property>\n')
        f.write(f'  <Property Name="Size">{20 + (10 + 240) * NoWidth}, {40 + 60 * NoHeight}</Property>\n')
        f.write(' </Object>\n')
        f.write('</Panel>\n')

# Helper: ErrCheckbox
def ErrCheckbox(msg, location, errtype, idx, sysvar_type):
    symbol_map = {
        "TO": "i_TOError",
        "CHK": "i_CHKError",
        "MC": "i_MCError",
        "DLC": "i_DLCError"
    }
    # Special case for CHK
    if errtype == "CHK":
        symbol = "i_CRC_CHKError"
    else:
        symbol = symbol_map.get(errtype, "i_TOError")
    return f'    <Object Type="Vector.CANalyzer.Panels.Design.CheckBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="CheckBoxControl{errtype}{idx}" Children="Controls" ControlName="Check Box">\n' \
           f'      <Property Name="Name">CheckBoxControl{errtype}{idx}</Property>\n' \
           f'      <Property Name="Size">50, 20</Property>\n' \
           f'      <Property Name="Location">{location}</Property>\n' \
           f'      <Property Name="Text">{errtype}</Property>\n' \
           f'      <Property Name="SymbolConfiguration">4;16;Error_panel_{sysvar_type};;;{symbol};1;;;{idx}</Property>\n' \
           f'    </Object>'

# Helper: ErrTextbox
def ErrTextbox(msg, location, errtype, idx, sysvar_type):
    symbol_map = {
        "TO": "i_TOErrorCounter",
        "CHK": "i_CHKErrorCounter",
        "MC": "i_MCErrorCounter",
        "DLC": "i_DLCErrorCounter"
    }
    symbol = symbol_map.get(errtype, "i_TOErrorCounter")
    # Name: TextBoxControl<errtype>Counter<idx>
    # Size: 20, 20
    # ValueDisplay: Decimal
    # DescriptionText: (empty)
    # DescriptionSize: 20, 20
    # DisplayLabel: Hide
    # SymbolConfiguration: 4;16;Error_panel_<sysvar_type>;;;symbol;1;;;idx
    return f'    <Object Type="Vector.CANalyzer.Panels.Design.TextBoxControl, Vector.CANalyzer.Panels.CommonControls, Version=8.0.35.0, Culture=neutral, PublicKeyToken=null" Name="TextBoxControl{errtype}Counter{idx}" Children="Controls" ControlName="Input/Output Box">\n' \
           f'      <Property Name="Name">TextBoxControl{errtype}Counter{idx}</Property>\n' \
           f'      <Property Name="Size">20, 20</Property>\n' \
           f'      <Property Name="Location">{location}</Property>\n' \
           f'      <Property Name="ValueDisplay">Decimal</Property>\n' \
           f'      <Property Name="DescriptionText"></Property>\n' \
           f'      <Property Name="DescriptionSize">20, 20</Property>\n' \
           f'      <Property Name="DisplayLabel">Hide</Property>\n' \
           f'      <Property Name="SymbolConfiguration">4;16;Error_panel_{sysvar_type};;;{symbol};1;;;{idx}</Property>\n' \
           f'    </Object>'


# --- Hàm liệt kê PUB messages theo loại E2E ---




# Hàm trả về string code C: hằng số, struct, mảng struct cho tất cả message TX

# Hàm sinh code C struct/mảng cho tất cả message TX, cho phép truyền mapping e2e_type
# Hàm trả về list các tuple (msg_name, e2e_type, e2e_cfg) cho tất cả message TX
def get_pub_messages_e2e_info(e2e_type_map=None, e2e_cfg_map=None):
    """
    Trả về list các tuple (msg_name, e2e_type, e2e_cfg) cho tất cả message TX.
    e2e_type mặc định E2E_NONE, e2e_cfg mặc định MSG_E2E_CFG_NULL.
    Cho phép truyền vào dict mapping từng message sang e2e_type/e2e_cfg nếu muốn.
    """
    working_folder = WorkingFolder_var.get().strip()
    if not working_folder.endswith("\\"):
        working_folder += "\\"
    excel_path = os.path.join(working_folder, "DBC_output.xlsx")
    if not os.path.isfile(excel_path):
        return []

    # Ưu tiên lấy e2e_type_map và e2e_cfg_map từ file path trong E2EConfigPath_var nếu có
    global E2EConfigPath_var
    config_path = E2EConfigPath_var.get().strip()
    if (not e2e_type_map or not e2e_cfg_map) and config_path and os.path.isfile(config_path):
        df_cfg = pd.read_csv(config_path)
        e2e_type_map = {str(row["MessageName"]).strip(): str(row["E2E_Type"]).strip() for _, row in df_cfg.iterrows()}
        e2e_cfg_map = {str(row["MessageName"]).strip(): str(row["E2E_Config"]).strip() for _, row in df_cfg.iterrows()}

    df = pd.read_excel(excel_path, sheet_name="Output_dbc")
    pub_msgs = df[(df["DIR"] == "TX")]
    if pub_msgs.empty:
        return []

    seen = set()
    result = []
    for _, row in pub_msgs.iterrows():
        msg_name = str(row["Message"]).strip()
        if not msg_name or msg_name in seen:
            continue
        seen.add(msg_name)
        # Ưu tiên lấy từ e2e_project_config_template.csv nếu có
        e2e_type = 'E2E_NONE'
        e2e_cfg = 'MSG_E2E_CFG_NULL'
        if e2e_type_map and msg_name in e2e_type_map:
            val = e2e_type_map[msg_name]
            if isinstance(val, int):
                if val == 1:
                    e2e_type = 'E2E_CRC8'
                elif val == 2:
                    e2e_type = 'E2E_XOR'
                elif val == 3:
                    e2e_type = 'E2E_CRC8_MULT'
                else:
                    e2e_type = 'E2E_NONE'
            elif isinstance(val, str):
                e2e_type = val if val in ['E2E_NONE','E2E_CRC8','E2E_XOR','E2E_CRC8_MULT'] else 'E2E_NONE'
        if e2e_cfg_map and msg_name in e2e_cfg_map:
            e2e_cfg = e2e_cfg_map[msg_name]
        result.append((msg_name, e2e_type, e2e_cfg))
    return result

PROJ_MSG_E2E_TX_ARR = []  # Danh sách dict: mỗi dict có 'PROJ_E2E_TYPE', 'PROJ_E2E_CFG', 'MSG_NAME'

# Hàm khởi tạo mảng E2E config cho handler Python
def init_proj_msg_e2e_tx_arr():
    global PROJ_MSG_E2E_TX_ARR
    e2e_list = get_pub_messages_e2e_info()
    PROJ_MSG_E2E_TX_ARR = []
    for msg_name, e2e_type, e2e_cfg in e2e_list:
        PROJ_MSG_E2E_TX_ARR.append({
            'MSG_NAME': msg_name,
            'PROJ_E2E_TYPE': e2e_type,
            'PROJ_E2E_CFG': e2e_cfg
        })

# ==== E2E Handler tổng quát cho PUB (Python version) ====
def PROJ_Sim_E2E_Handler(MSGIndex, MSG):
    """
    Hàm tổng quát gọi rollingcounter và checksum theo E2E_TYPE
    MSGIndex: int
    MSG: dict hoặc object message
    """
    e2e_type = PROJ_MSG_E2E_TX_ARR[MSGIndex]['PROJ_E2E_TYPE']
    if e2e_type == 'E2E_XOR':
        PROJ_Sim_Rollingcounter_XOR(MSGIndex, MSG)
        PROJ_Sim_Checksum_XOR(MSGIndex, MSG)
    elif e2e_type == 'E2E_CRC8':
        PROJ_Sim_Rollingcounter_CRC8(MSGIndex, MSG)
        PROJ_Sim_Checksum_CRC8(MSGIndex, MSG)
    elif e2e_type == 'E2E_CRC8_MULT':
        PROJ_Sim_Rollingcounter_CRC8_MULT(MSGIndex, MSG)
        PROJ_Sim_Checksum_CRC8_MULT(MSGIndex, MSG)
    else:
        # Không xử lý
        pass

# ==== Block xử lý cho từng loại E2E (chỉ là khung, bạn copy logic chi tiết vào từng hàm này) ====
def PROJ_Sim_Rollingcounter_XOR(MSGIndex, MSG):
    """Logic rollingcounter cho E2E_XOR (PUB)"""
    # Mapping MSGIndex sang tên message
    msg_name = PROJ_MSG_E2E_TX_ARR[MSGIndex]['MSG_NAME']
    # Xác định t_ALC_position
    if msg_name in [
        'ABS_ESP_1', 'EMS_1']:
        t_ALC_position = 7
    elif msg_name in [
        'ABS_ESP_3', 'ABS_ESP_4', 'ABS_ESP_7', 'ABS_ESP_8', 'BCM_1', 'ICM_5', 'TCU_2', 'EMS_3']:
        t_ALC_position = 6
    elif msg_name in ['YAS_1', 'YAS_2']:
        t_ALC_position = 0
    else:
        t_ALC_position = 3

    # Xác định t_ALC_Offset
    if msg_name in [
        'ABS_ESP_1', 'ABS_ESP_3', 'ABS_ESP_4', 'ABS_ESP_7', 'ABS_ESP_8', 'BCM_1', 'EMS_1', 'TCU_2', 'ICM_5']:
        t_ALC_Offset = 0
    elif msg_name in ['YAS_1', 'YAS_2']:
        t_ALC_Offset = 1
    else:
        t_ALC_Offset = 4

    # Lấy rolling counter hiện tại
    t_LastSeqCounter = 0
    t_LastSeqCounter_byte = 0
    if t_ALC_Offset == 0:
        t_LastSeqCounter = MSG['data'][t_ALC_position] & 0x0F
    elif t_ALC_Offset == 4:
        t_LastSeqCounter = (MSG['data'][t_ALC_position] & 0xF0) >> 4
    elif t_ALC_Offset == 1:
        t_LastSeqCounter_byte = MSG['data'][0]
    else:
        t_LastSeqCounter = (MSG['data'][t_ALC_position] & 0xF0) >> 4

    # Giả lập các biến lỗi (bạn cần thay bằng biến thực tế nếu có)
    Error_panel = globals().get('Error_panel_PUB', {})
    TOError = Error_panel.get('i_TOError', [0]*256)
    CRC_CHKError = Error_panel.get('i_CRC_CHKError', [0]*256)
    MCError = Error_panel.get('i_MCError', [1]*256)
    MCErrorCounter = Error_panel.get('i_MCErrorCounter', [0]*256)

    # Rolling Counter logic
    if (TOError[MSGIndex] == 0 and CRC_CHKError[MSGIndex] == 0 and MCError[MSGIndex]):
        if MCErrorCounter[MSGIndex] > 0:
            MCErrorCounter[MSGIndex] -= 1
        if MCErrorCounter[MSGIndex] == 0:
            MCError[MSGIndex] = 0
            MCErrorCounter[MSGIndex] = -1
    else:
        t_LastSeqCounter = (t_LastSeqCounter + 1) % 16
        t_LastSeqCounter_byte = (t_LastSeqCounter_byte + 1) % 256

    # Gán lại giá trị rolling counter vào MSG
    if t_ALC_Offset == 0:
        MSG['data'][t_ALC_position] = (MSG['data'][t_ALC_position] & 0xF0) | (t_LastSeqCounter & 0x0F)
    elif t_ALC_Offset == 4:
        MSG['data'][t_ALC_position] = (MSG['data'][t_ALC_position] & 0x0F) | ((t_LastSeqCounter & 0x0F) << 4)
    elif t_ALC_Offset == 1:
        MSG['data'][0] = t_LastSeqCounter_byte & 0xFF
    else:
        MSG['data'][t_ALC_position] = (MSG['data'][t_ALC_position] & 0x0F) | ((t_LastSeqCounter & 0x0F) << 4)

def PROJ_Sim_Checksum_XOR(MSGIndex, MSG):
    """Copy logic checksum cho E2E_XOR ở đây"""
    pass

def PROJ_Sim_Rollingcounter_CRC8(MSGIndex, MSG):
    """Copy logic rollingcounter cho E2E_CRC8 ở đây"""
    pass

def PROJ_Sim_Checksum_CRC8(MSGIndex, MSG):
    """Copy logic checksum cho E2E_CRC8 ở đây"""
    pass

def PROJ_Sim_Rollingcounter_CRC8_MULT(MSGIndex, MSG):
    """Copy logic rollingcounter cho E2E_CRC8_MULT ở đây"""
    pass

def PROJ_Sim_Checksum_CRC8_MULT(MSGIndex, MSG):
    """Copy logic checksum cho E2E_CRC8_MULT ở đây"""
    pass



def gen_e2e_project_config():
    """Sinh file e2e_project_config_template.csv từ DBC_output.xlsx (sheet Output_dbc)"""
    working_folder = WorkingFolder_var.get().strip()
    if not working_folder.endswith("\\"): working_folder += "\\"
    excel_path = os.path.join(working_folder, "DBC_output.xlsx")
    if not os.path.isfile(excel_path):
        messagebox.showerror("Lỗi", f"Không tìm thấy file {excel_path}")
        return
    df = pd.read_excel(excel_path, sheet_name="Output_dbc")
    pub_msgs = df[df["DIR"] == "TX"]
    seen = set()
    rows = []
    for _, row in pub_msgs.iterrows():
        msg_name = str(row["Message"]).strip()
        if not msg_name or msg_name in seen:
            continue
        seen.add(msg_name)
        rows.append({"MessageName": msg_name, "E2E_Type": "E2E_NONE", "E2E_Config": "MSG_E2E_CFG_NULL"})
    out_path = os.path.join(working_folder, "e2e_project_config_template.csv")
    pd.DataFrame(rows).to_csv(out_path, index=False)
    E2EConfigPath_var.set(out_path)
    save_log()
    messagebox.showinfo("Gen e2e_project_config", f"Đã sinh file: {out_path}")

def load_e2e_project_config():
    """Cho user chọn file e2e_project_config_template.csv, nạp vào e2e_type_map và e2e_cfg_map"""
    global e2e_type_map, e2e_cfg_map
    file_path = filedialog.askopenfilename(title="Chọn file e2e_project_config", filetypes=[("CSV files", "*.csv")])
    if not file_path: return
    df = pd.read_csv(file_path)
    e2e_type_map = {str(row["MessageName"]).strip(): str(row["E2E_Type"]).strip() for _, row in df.iterrows()}
    e2e_cfg_map = {str(row["MessageName"]).strip(): str(row["E2E_Config"]).strip() for _, row in df.iterrows()}
    E2EConfigPath_var.set(file_path)
    save_log()
    messagebox.showinfo("Load e2e_project_config", f"Đã load file: {file_path}\nSố message: {len(e2e_type_map)}")


# Nút chức năng

def AutoGen_click():
    # Lấy thông tin DBC từ GUI
    # Lấy giá trị DBC path và Channel từ GUI
    dbc_path = DBC_var.get()
    channel = Channel_var.get().strip().upper()
    # Nếu Channel là PRI thì chỉ thay DBC private, nếu PUB thì chỉ thay DBC public
    if channel == 'PRI':
        Autogen.autogen_create_cfg('', dbc_path, 'PRI')
    elif channel == 'PUB':
        Autogen.autogen_create_cfg(dbc_path, '', 'PUB')
    else:
        from tkinter import messagebox
        messagebox.showinfo('AutoGen', 'Vui lòng chọn Channel là PUB hoặc PRI!')


# Thêm 2 nút mới cho E2E Project Config




import tkinter.ttk as ttk
root = tk.Tk()
root.title('RestBus AIO')
root.geometry('650x260')

# Biến lưu path file e2e_project_config_template.csv cho GUI
E2EConfigPath_var = tk.StringVar()
# Đặt icon cho phần mềm
icon_path = r'D:\05_Restbus_AIO\icon\icon.png'
if os.path.exists(icon_path):
    try:
        icon_img = tk.PhotoImage(file=icon_path)
        root.iconphoto(True, icon_img)
    except Exception:
        pass

# Biến giao diện
WorkingFolder_var = tk.StringVar()
DBC_var = tk.StringVar()
Node_var = tk.StringVar()
Channel_var = tk.StringVar()
tb_RestbusPath_var = tk.StringVar()

# Load log nếu có
log_vals = load_log()
WorkingFolder_var.set(log_vals[0])
DBC_var.set(log_vals[1])
Node_var.set(log_vals[2])
Channel_var.set(log_vals[3])
tb_RestbusPath_var.set(log_vals[4])
E2EConfigPath_var.set(log_vals[5])

# Liên kết sự kiện thay đổi
WorkingFolder_var.trace_add('write', workingfolder_change)
DBC_var.trace_add('write', dbc_change)
Node_var.trace_add('write', node_change)
Channel_var.trace_add('write', channel_change)
tb_RestbusPath_var.trace_add('write', tb_RestbusPath_change)

# Main frame
mainframe = ttk.Frame(root, padding="10 10 10 10")
mainframe.grid(row=0, column=0, sticky="nsew")
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

labels = [
    # Đẩy các label xuống 1 dòng để Project ở trên cùng
    ('Working Folder:', WorkingFolder_var, bt_curdir_click),
    ('DBC Path:', DBC_var, bt_opendbc_click),
    ('Node Name:', Node_var, None),
    ('Channel:', Channel_var, None),
    ('Restbus Path:', tb_RestbusPath_var, None),
]
# Project droplist
Project_var = tk.StringVar()
Project_var.set('NA')
# Thêm Project droplist vào giao diện
project_label = ttk.Label(mainframe, text="Project" )
project_label.grid(row=0, column=0, sticky='w', padx=3, pady=3)
project_options = ['NA', 'Cherry', 'ProjectB']  # Cập nhật danh sách project thực tế tại đây
project_combo = ttk.Combobox(mainframe, textvariable=Project_var, values=project_options, state='readonly')
project_combo.grid(row=0, column=1, sticky='ew', padx=3, pady=3)


for i, (label, var, btn) in enumerate(labels):
    ttk.Label(mainframe, text=label).grid(row=i+1, column=0, sticky='e', padx=3, pady=3)
    if label == 'Channel:':
        channel_combobox = ttk.Combobox(mainframe, textvariable=Channel_var, values=["PUB", "PRI"], state="readonly")
        channel_combobox.grid(row=i+1, column=1, sticky='ew', padx=3, pady=3)
        channel_combobox.set("PUB")
    else:
        ttk.Entry(mainframe, textvariable=var).grid(row=i+1, column=1, sticky='ew', padx=3, pady=3)
    # Hiển thị đúng nút chức năng cho từng dòng
    if label == 'Working Folder:':
        ttk.Button(mainframe, text='Current Dir', width=12, command=bt_curdir_click).grid(row=i, column=2, sticky='w', padx=3, pady=3)
    elif label == 'DBC Path:':
        ttk.Button(mainframe, text='Open DBC', width=12, command=bt_opendbc_click).grid(row=i, column=2, sticky='w', padx=3, pady=3)

# Hàng 1: các nút chính
btns_main = [
    ('Get DBC Info', GetDBCInfor_Button_Click),
    ('Open Excel Output', open_dbc_output),

]
for i, (text, cmd) in enumerate(btns_main):
    ttk.Button(mainframe, text=text, command=cmd).grid(row=len(labels), column=i, sticky='ew', padx=5, pady=10)

# Hàng 2: các nút E2E config
btns_e2e = [
    ('Gen_E2E_config', gen_e2e_project_config),
    ('Load_E2E_config', load_e2e_project_config),
    ('Generate', GenerateButton_click),
    # ('AutoGen', AutoGen_click),
]

# Thêm các nút E2E config
for i, (text, cmd) in enumerate(btns_e2e):
    ttk.Button(mainframe, text=text, command=cmd).grid(row=len(labels)+1, column=i, sticky='ew', padx=5, pady=10)

# Thêm ô hiển thị path file e2e_project_config_template.csv ngay dưới hàng nút E2E config
ttk.Label(mainframe, text='E2E Config Path:').grid(row=len(labels)+2, column=0, sticky='e', padx=3, pady=3)
ttk.Entry(mainframe, textvariable=E2EConfigPath_var, state='readonly').grid(row=len(labels)+2, column=1, columnspan=3, sticky='ew', padx=3, pady=3)

# Cấu hình co giãn
for i in range(len(labels)):
    mainframe.rowconfigure(i, weight=1)
mainframe.rowconfigure(len(labels), weight=1)
mainframe.columnconfigure(1, weight=1)

root.minsize(500, 220)
root.mainloop()
